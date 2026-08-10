#!/usr/bin/env python3
"""Generate Scenario Performance and Robustness.

Plan: Evaluate the complete 162-combination primary factorial across affected-
population bounds, per-capita water demand, fleet size, road state, and point
state. Framework: AnaSOP Sections 5-7 full-mesh demand accounting, the 2,000 m
extended allocation diagnostic, route-feasible tanker trip and work-time
capacity, road-edge failure states, and a highest-load single announced-point
failure screen. Each row is a deterministic full-mesh feasibility allocation:
reachable meshes are ordered by shortest assigned access and admitted whole
until route-feasible fleet capacity is exhausted. The table is a robustness
screen, not a turn-by-turn dispatch schedule or observed 2026 performance.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import textwrap

import geopandas as gpd
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont

import figure_scenario_based_tanker_and_temporary_water_point_allocation as allocation


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "data/results/tables/Table_scenario_performance_and_robustness.xlsx"
PREVIEW = ROOT / "data/exp/Table_scenario_performance_and_robustness.png"
APPENDIX_OUTPUT = (
    ROOT
    / "data/results/tables/Table_full_factorial_scenario_performance_and_robustness_appendix.xlsx"
)
APPENDIX_PREVIEW = (
    ROOT
    / "data/exp/appendix/Appendix_full_factorial_scenario_performance_and_robustness.png"
)

TITLE = "Scenario Performance and Robustness"
SHEET_NAME = "Robustness Summary"
APPENDIX_TITLE = "Full-Factorial Scenario Performance and Robustness Appendix"
APPENDIX_SHEET_NAME = "Full Factorial"
ACCESS_DISTANCE_M = 2_000.0
FLEET_SIZES = [5, 10, 20]

OUTAGE_SCENARIOS = {
    "lower_one_person_per_household": "Lower",
    "proportional_central": "Central",
    "upper_p90_household_size": "Upper",
}
DEMAND_SCENARIOS = {
    "minimum": "Minimum (3 L/person/day)",
    "basic": "Basic (10 L/person/day)",
    "extended": "Extended (20 L/person/day)",
}
ROAD_SCENARIOS = {
    "baseline": "Baseline",
    "matched_restrictions_closed": "Matched Restrictions Closed",
    "severe_disruption": "Severe Disruption",
}
POINT_SCENARIOS = [
    "Reported Schedule",
    "Highest-Load Single-Point Failure",
]

HEADERS = [
    "Outage Population",
    "Demand Standard",
    "Fleet Size",
    "Road State",
    "Water-Point State",
    "Protected Population Share",
    "Unmet Water (L/day)",
    "Mean Assigned Access (m)",
    "Sites Used",
    "Route Feasible",
]
MAIN_HEADERS = [
    "Outage Population",
    "Demand Standard",
    "Fleet Size",
    "Baseline Protected Share",
    "Baseline Unmet Water (L/day)",
    "Baseline Mean Access (m)",
    "High-Load Point-Failure Loss (pp)",
    "Matched-Road Restriction Loss (pp)",
    "Combined High-Load + Road Stress Protected Share",
]
PREVIEW_HEADERS = [
    "Outage\nPopulation",
    "Demand\nStandard",
    "Fleet\nSize",
    "Road State",
    "Water-Point State",
    "Protected\nPopulation Share",
    "Unmet Water\n(L/day)",
    "Mean Assigned\nAccess (m)",
    "Sites\nUsed",
    "Route\nFeasible",
]
MAIN_PREVIEW_HEADERS = [
    "Outage\nPopulation",
    "Demand\nStandard",
    "Fleet\nSize",
    "Baseline\nProtected Share",
    "Baseline Unmet\nWater (L/day)",
    "Baseline Mean\nAccess (m)",
    "High-Load Point-Failure\nLoss (pp)",
    "Matched-Road Restriction\nLoss (pp)",
    "Combined High-Load + Road\nStress Protected Share",
]

NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"
JAPANESE_PATTERN = re.compile(r"[\u3040-\u30ff\u3400-\u9fff]")


@dataclass
class StructuralScenario:
    road_key: str
    point_label: str
    scenario: allocation.ScenarioInputs | None
    failed_point: str | None


def load_inputs() -> tuple[
    pd.DataFrame,
    gpd.GeoDataFrame,
    pd.DataFrame,
    gpd.GeoDataFrame,
    pd.DataFrame,
    gpd.GeoDataFrame,
    gpd.GeoDataFrame,
]:
    parameters = pd.read_parquet(allocation.PARAMETERS)
    nodes = gpd.read_parquet(
        allocation.ROAD_NODES, columns=["Network Node ID", "Geometry"]
    )
    edges = pd.read_parquet(
        allocation.ROAD_EDGES,
        columns=[
            "Road Edge ID",
            "From Node ID",
            "To Node ID",
            "Road Length (m)",
            "Baseline Edge Travel Time (min)",
            "Hazard Exposure Class",
            "Road Available",
            "Network Analysis Eligible",
        ],
    )
    demand = gpd.read_parquet(allocation.MESH_DEMAND)
    access = pd.read_parquet(allocation.MESH_ACCESS)
    water = gpd.read_parquet(allocation.WATER_POINTS)
    staging = gpd.read_parquet(allocation.STAGING)
    return parameters, nodes, edges, demand, access, water, staging


def road_closure_sets(edges: pd.DataFrame) -> dict[str, set[str]]:
    restrictions = pd.read_parquet(
        allocation.RESTRICTIONS,
        columns=["Matched Road Edge ID", "Road Edge Match Status"],
    )
    matched = set(
        restrictions.loc[
            restrictions["Road Edge Match Status"]
            .astype("string")
            .str.startswith("matched", na=False),
            "Matched Road Edge ID",
        ]
        .dropna()
        .astype(str)
    )
    hazard = set(
        edges.loc[edges["Hazard Exposure Class"].notna(), "Road Edge ID"].astype(str)
    )
    return {
        "baseline": set(),
        "matched_restrictions_closed": matched,
        "severe_disruption": matched | hazard,
    }


def highest_load_failure(
    scenario: allocation.ScenarioInputs,
) -> tuple[str, allocation.ScenarioInputs]:
    announced = np.flatnonzero(~scenario.sites["Temporary Site"].to_numpy(bool))
    announced_arcs = scenario.arcs.loc[
        scenario.arcs["Site Index"].isin(announced)
    ].sort_values(["Unit Index", "Distance (m)", "Site Index"], kind="stable")
    if announced_arcs.empty:
        raise ValueError(f"No announced point remains in {scenario.label}")
    nearest = announced_arcs.drop_duplicates("Unit Index", keep="first")
    demand = scenario.units["Estimated Water Demand (L/day)"].to_numpy(float)
    load = np.bincount(
        nearest["Site Index"].to_numpy(np.int32),
        weights=demand[nearest["Unit Index"].to_numpy(np.int32)],
        minlength=len(scenario.sites),
    )
    failed_index = int(announced[np.argmax(load[announced])])
    failed_name = str(scenario.sites.iloc[failed_index]["Site Name"])
    return failed_name, allocation.remove_announced_site(scenario, failed_index)


def prepare_structural_scenarios() -> tuple[
    list[StructuralScenario], gpd.GeoDataFrame
]:
    parameters, nodes, edges, demand, access, water, staging = load_inputs()
    closures = road_closure_sets(edges)
    older_weight = allocation.older_priority_weights()
    structural: list[StructuralScenario] = []
    for road_key, closed_edges in closures.items():
        try:
            reported = allocation.prepare_scenario(
                ROAD_SCENARIOS[road_key],
                closed_edges,
                None,
                nodes,
                edges.drop(columns=["Hazard Exposure Class"]),
                demand,
                access,
                water,
                staging,
                parameters,
                older_weight,
                access_distance_m=ACCESS_DISTANCE_M,
            )
        except (ValueError, RuntimeError) as error:
            print(f"Structural scenario infeasible: {road_key}: {error}")
            structural.extend(
                [
                    StructuralScenario(road_key, POINT_SCENARIOS[0], None, None),
                    StructuralScenario(road_key, POINT_SCENARIOS[1], None, None),
                ]
            )
            continue
        structural.append(
            StructuralScenario(road_key, POINT_SCENARIOS[0], reported, None)
        )
        failed_name, failed = highest_load_failure(reported)
        structural.append(
            StructuralScenario(road_key, POINT_SCENARIOS[1], failed, failed_name)
        )
        print(
            f"Prepared {ROAD_SCENARIOS[road_key]}: "
            f"closed edges={len(closed_edges):,}; "
            f"highest-load failed point={failed_name}"
        )
    return structural, demand


def route_capacity_per_vehicle(scenario: allocation.ScenarioInputs) -> int:
    active_sites = set(scenario.arcs["Site Index"].astype(int).unique())
    best = 0
    for route_row, site_index in enumerate(scenario.route_site_index):
        if int(site_index) not in active_sites:
            continue
        refill_index = int(scenario.route_refill_index[route_row])
        available = scenario.work_minutes - scenario.refill_base_minutes[refill_index]
        cycle = float(scenario.route_cycle_minutes[route_row])
        trips = min(
            scenario.trip_limit,
            max(0, int(np.floor(available / cycle))) if cycle > 0 else 0,
        )
        best = max(best, trips)
    return best


def demand_vectors(
    demand: pd.DataFrame,
    units: pd.DataFrame,
    outage_key: str,
    demand_key: str,
) -> tuple[np.ndarray, np.ndarray]:
    selected = demand.loc[
        demand["Outage Population Scenario"].eq(outage_key)
        & demand["Demand Scenario"].eq(demand_key),
        ["Mesh Code", "Estimated Outage Population", "Estimated Water Demand (L/day)"],
    ].copy()
    selected["Mesh Code"] = selected["Mesh Code"].astype(str)
    selected = selected.drop_duplicates("Mesh Code").set_index("Mesh Code")
    ordered = selected.reindex(units["Mesh Code"].astype(str))
    if ordered.isna().any().any():
        raise ValueError(f"Demand vector is incomplete for {outage_key}/{demand_key}")
    return (
        ordered["Estimated Outage Population"].to_numpy(float),
        ordered["Estimated Water Demand (L/day)"].to_numpy(float),
    )


def feasible_mesh_allocation(
    scenario: allocation.ScenarioInputs,
    population: np.ndarray,
    water: np.ndarray,
    fleet_size: int,
) -> dict[str, float | int | str]:
    nearest = scenario.arcs.sort_values(
        ["Unit Index", "Distance (m)", "Site Index"], kind="stable"
    ).drop_duplicates("Unit Index", keep="first")
    unit_index = nearest["Unit Index"].to_numpy(np.int32)
    positive = water[unit_index] > 0
    nearest = nearest.loc[positive].copy()
    unit_index = nearest["Unit Index"].to_numpy(np.int32)
    trips_per_vehicle = route_capacity_per_vehicle(scenario)
    route_feasible = trips_per_vehicle > 0 and len(nearest) > 0
    if not route_feasible:
        return {
            "Protected Population Share": 0.0,
            "Unmet Water (L/day)": float(water.sum()),
            "Mean Assigned Access (m)": np.nan,
            "Sites Used": 0,
            "Route Feasible": "No",
        }

    capacity = (
        float(fleet_size)
        * float(trips_per_vehicle)
        * float(scenario.trip_capacity_liters)
    )
    nearest["Population"] = population[unit_index]
    nearest["Water"] = water[unit_index]
    nearest = nearest.sort_values(
        ["Distance (m)", "Water", "Unit Index", "Site Index"], kind="stable"
    )
    remaining = capacity
    selected_rows: list[int] = []
    for row_index, row in nearest.iterrows():
        required = float(row["Water"])
        if required <= remaining + 1e-9:
            selected_rows.append(row_index)
            remaining -= required
        if remaining <= 1e-9:
            break
    selected = nearest.loc[selected_rows]
    protected_population = float(selected["Population"].sum())
    delivered_water = float(selected["Water"].sum())
    total_population = float(population.sum())
    total_water = float(water.sum())
    mean_access = (
        float(
            np.average(
                selected["Distance (m)"].to_numpy(float),
                weights=selected["Population"].to_numpy(float),
            )
        )
        if protected_population > 0
        else np.nan
    )
    return {
        "Protected Population Share": (
            protected_population / total_population if total_population > 0 else np.nan
        ),
        "Unmet Water (L/day)": max(0.0, total_water - delivered_water),
        "Mean Assigned Access (m)": mean_access,
        "Sites Used": int(selected["Site Index"].nunique()),
        "Route Feasible": "Yes",
    }


def construct_table() -> tuple[pd.DataFrame, dict[str, object]]:
    structural, demand = prepare_structural_scenarios()
    rows: list[dict[str, object]] = []
    failure_names: dict[str, str] = {}
    for structure in structural:
        if structure.failed_point is not None:
            failure_names[structure.road_key] = structure.failed_point
        for outage_key, outage_label in OUTAGE_SCENARIOS.items():
            for demand_key, demand_label in DEMAND_SCENARIOS.items():
                for fleet_size in FLEET_SIZES:
                    if structure.scenario is None:
                        scenario_demand = demand.loc[
                            demand["Outage Population Scenario"].eq(outage_key)
                            & demand["Demand Scenario"].eq(demand_key),
                            "Estimated Water Demand (L/day)",
                        ]
                        result = {
                            "Protected Population Share": 0.0,
                            "Unmet Water (L/day)": float(scenario_demand.sum()),
                            "Mean Assigned Access (m)": np.nan,
                            "Sites Used": 0,
                            "Route Feasible": "No",
                        }
                    else:
                        population, water = demand_vectors(
                            demand,
                            structure.scenario.units,
                            outage_key,
                            demand_key,
                        )
                        result = feasible_mesh_allocation(
                            structure.scenario, population, water, fleet_size
                        )
                    rows.append(
                        {
                            "Outage Population": outage_label,
                            "Demand Standard": demand_label,
                            "Fleet Size": fleet_size,
                            "Road State": ROAD_SCENARIOS[structure.road_key],
                            "Water-Point State": structure.point_label,
                            **result,
                        }
                    )
    table_data = pd.DataFrame(rows, columns=HEADERS)
    road_order = {label: order for order, label in enumerate(ROAD_SCENARIOS.values())}
    point_order = {label: order for order, label in enumerate(POINT_SCENARIOS)}
    outage_order = {label: order for order, label in enumerate(OUTAGE_SCENARIOS.values())}
    demand_order = {label: order for order, label in enumerate(DEMAND_SCENARIOS.values())}
    table_data = table_data.sort_values(
        by=HEADERS[:5],
        key=lambda series: series.map(
            road_order
            if series.name == "Road State"
            else point_order
            if series.name == "Water-Point State"
            else outage_order
            if series.name == "Outage Population"
            else demand_order
            if series.name == "Demand Standard"
            else {value: index for index, value in enumerate(FLEET_SIZES)}
        ),
        kind="stable",
    ).reset_index(drop=True)
    audit = {"failure_names": failure_names}
    return table_data, audit


def summarize_for_main_text(full_table: pd.DataFrame) -> pd.DataFrame:
    keys = ["Outage Population", "Demand Standard", "Fleet Size"]

    def select(road_state: str, point_state: str, prefix: str) -> pd.DataFrame:
        selected = full_table.loc[
            full_table["Road State"].eq(road_state)
            & full_table["Water-Point State"].eq(point_state),
            keys
            + [
                "Protected Population Share",
                "Unmet Water (L/day)",
                "Mean Assigned Access (m)",
            ],
        ].copy()
        return selected.rename(
            columns={
                "Protected Population Share": f"{prefix} Protected Share",
                "Unmet Water (L/day)": f"{prefix} Unmet Water",
                "Mean Assigned Access (m)": f"{prefix} Mean Access",
            }
        )

    baseline = select("Baseline", "Reported Schedule", "Baseline")
    point_failure = select(
        "Baseline", "Highest-Load Single-Point Failure", "Point Failure"
    )
    road_restriction = select(
        "Matched Restrictions Closed", "Reported Schedule", "Road Restriction"
    )
    combined = select(
        "Matched Restrictions Closed",
        "Highest-Load Single-Point Failure",
        "Combined Stress",
    )
    summary = baseline.merge(point_failure, on=keys, validate="one_to_one")
    summary = summary.merge(road_restriction, on=keys, validate="one_to_one")
    summary = summary.merge(combined, on=keys, validate="one_to_one")
    summary["High-Load Point-Failure Loss (pp)"] = (
        100.0
        * (
            summary["Baseline Protected Share"]
            - summary["Point Failure Protected Share"]
        )
    ).clip(lower=0.0)
    summary["Matched-Road Restriction Loss (pp)"] = (
        100.0
        * (
            summary["Baseline Protected Share"]
            - summary["Road Restriction Protected Share"]
        )
    ).clip(lower=0.0)
    summary["Combined Stress Protected Share"] = np.minimum(
        summary["Combined Stress Protected Share"],
        summary["Baseline Protected Share"],
    )
    summary = summary.rename(
        columns={
            "Baseline Unmet Water": "Baseline Unmet Water (L/day)",
            "Baseline Mean Access": "Baseline Mean Access (m)",
            "Combined Stress Protected Share": (
                "Combined High-Load + Road Stress Protected Share"
            ),
        }
    )
    return summary[MAIN_HEADERS].reset_index(drop=True)


def validate_table(table_data: pd.DataFrame) -> None:
    assert table_data.columns.tolist() == HEADERS
    assert table_data.shape == (162, 10), table_data.shape
    assert set(table_data["Outage Population"]) == set(OUTAGE_SCENARIOS.values())
    assert set(table_data["Demand Standard"]) == set(DEMAND_SCENARIOS.values())
    assert set(table_data["Fleet Size"]) == set(FLEET_SIZES)
    assert set(table_data["Road State"]) == set(ROAD_SCENARIOS.values())
    assert set(table_data["Water-Point State"]) == set(POINT_SCENARIOS)
    assert table_data["Protected Population Share"].between(0, 1).all()
    assert table_data["Unmet Water (L/day)"].ge(0).all()
    assert table_data["Sites Used"].ge(0).all()
    assert set(table_data["Route Feasible"]).issubset({"Yes", "No"})
    combination_count = table_data.groupby(HEADERS[:5], observed=True).size()
    assert len(combination_count) == 162 and combination_count.eq(1).all()
    monotonic_groups = [
        "Outage Population", "Demand Standard", "Road State", "Water-Point State"
    ]
    for _, group in table_data.groupby(monotonic_groups, observed=True):
        ordered = group.sort_values("Fleet Size")
        assert np.all(np.diff(ordered["Protected Population Share"]) >= -1e-9)
        assert np.all(np.diff(ordered["Unmet Water (L/day)"]) <= 1e-6)
    for column in table_data.select_dtypes(include=["object", "string"]).columns:
        assert not table_data[column].astype(str).map(
            lambda value: bool(JAPANESE_PATTERN.search(value))
        ).any(), f"Japanese text remains in {column}"


def validate_main_table(table_data: pd.DataFrame) -> None:
    assert table_data.columns.tolist() == MAIN_HEADERS
    assert table_data.shape == (27, 9), table_data.shape
    assert table_data["Baseline Protected Share"].between(0, 1).all()
    assert table_data[
        "Combined High-Load + Road Stress Protected Share"
    ].between(0, 1).all()
    assert table_data["Baseline Unmet Water (L/day)"].ge(0).all()
    assert table_data["Baseline Mean Access (m)"].between(0, ACCESS_DISTANCE_M).all()
    assert table_data["High-Load Point-Failure Loss (pp)"].ge(-1e-8).all()
    assert table_data["Matched-Road Restriction Loss (pp)"].ge(-1e-8).all()
    assert (
        table_data["Combined High-Load + Road Stress Protected Share"]
        <= table_data["Baseline Protected Share"] + 1e-9
    ).all()
    for column in table_data.select_dtypes(include=["object", "string"]).columns:
        assert not table_data[column].astype(str).map(
            lambda value: bool(JAPANESE_PATTERN.search(value))
        ).any(), f"Japanese text remains in {column}"


def write_appendix_workbook(
    table_data: pd.DataFrame, audit: dict[str, object]
) -> None:
    APPENDIX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = APPENDIX_SHEET_NAME
    sheet.sheet_view.showGridLines = False
    # Preserve the title in row 1 without merged cells for deterministic DOCX export.
    sheet["A1"] = APPENDIX_TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 11):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 54

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(table_data.iterrows()):
        excel_row = offset + 3
        fill = PALE_BLUE if offset % 2 == 0 else WHITE
        for column, header in enumerate(HEADERS, start=1):
            value = record[header]
            if pd.isna(value):
                value = None
            cell = sheet.cell(excel_row, column, value)
            left = column in (1, 2, 4, 5, 10)
            cell.font = Font(
                name="Aptos", size=9.1,
                bold=column in (1, 4),
                color=NAVY if column in (1, 4) else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if left else "right",
                vertical="center", wrap_text=left,
            )
        sheet.cell(excel_row, 3).number_format = "0"
        sheet.cell(excel_row, 6).number_format = "0.0%"
        sheet.cell(excel_row, 7).number_format = "#,##0"
        sheet.cell(excel_row, 8).number_format = "#,##0.0"
        sheet.cell(excel_row, 9).number_format = "0"
        sheet.row_dimensions[excel_row].height = 29

    end_row = len(table_data) + 2
    excel_table = Table(displayName="ScenarioRobustness", ref=f"A2:J{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)
    sheet.conditional_formatting.add(
        f"F3:F{end_row}",
        ColorScaleRule(
            start_type="min", start_color="F4CCCC",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="D9EAD3",
        ),
    )
    widths = {
        "A": 20, "B": 27, "C": 12, "D": 28, "E": 35,
        "F": 27, "G": 24, "H": 27, "I": 14, "J": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "F3"
    sheet.auto_filter.ref = f"A2:J{end_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:J{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · 2,000 m extended allocation diagnostic · Full-mesh feasibility screen"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY
    workbook.properties.title = APPENDIX_TITLE
    workbook.properties.subject = "KE01d scenario performance and robustness"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = (
        "Kumamoto, emergency water, robustness, road disruption, point failure"
    )
    workbook.save(APPENDIX_OUTPUT)


def write_main_workbook(table_data: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False
    # Preserve the title in row 1 without merged cells for deterministic DOCX export.
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 10):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(MAIN_HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 58

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(table_data.iterrows()):
        excel_row = offset + 3
        fill = PALE_BLUE if offset % 2 == 0 else WHITE
        for column, header in enumerate(MAIN_HEADERS, start=1):
            cell = sheet.cell(excel_row, column, record[header])
            left = column in (1, 2)
            cell.font = Font(
                name="Aptos", size=9.2,
                bold=column == 1,
                color=NAVY if column == 1 else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if left else "right",
                vertical="center", wrap_text=left,
            )
        sheet.cell(excel_row, 3).number_format = "0"
        sheet.cell(excel_row, 4).number_format = "0.0%"
        sheet.cell(excel_row, 5).number_format = "#,##0"
        sheet.cell(excel_row, 6).number_format = "#,##0.0"
        sheet.cell(excel_row, 7).number_format = "0.0"
        sheet.cell(excel_row, 8).number_format = "0.0"
        sheet.cell(excel_row, 9).number_format = "0.0%"
        sheet.row_dimensions[excel_row].height = 32

    end_row = len(table_data) + 2
    excel_table = Table(displayName="ScenarioRobustnessSummary", ref=f"A2:I{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)
    for column in ("D", "I"):
        sheet.conditional_formatting.add(
            f"{column}3:{column}{end_row}",
            ColorScaleRule(
                start_type="min", start_color="F4CCCC",
                mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                end_type="max", end_color="D9EAD3",
            ),
        )
    widths = {
        "A": 20, "B": 27, "C": 12, "D": 25, "E": 26,
        "F": 25, "G": 24, "H": 27, "I": 31,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "D3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:I{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · Main-text summary · Full 162-scenario matrix reported in appendix"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY
    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d main-text scenario robustness summary"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = (
        "Kumamoto, emergency water, robustness, main-text summary"
    )
    workbook.save(OUTPUT)


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def format_preview_value(header: str, value: object) -> str:
    if pd.isna(value):
        return "—"
    if "Protected Share" in header or header == "Protected Population Share":
        return f"{float(value):.1%}"
    if "Unmet Water" in header:
        return f"{float(value):,.0f}"
    if "Mean Access" in header or header == "Mean Assigned Access (m)":
        return f"{float(value):,.1f}"
    if "Loss (pp)" in header:
        return f"{float(value):.1f}"
    if header in {"Fleet Size", "Sites Used"}:
        return f"{int(value):,}"
    return str(value)


def wrapped_lines(
    draw: ImageDraw.ImageDraw,
    value: str,
    chosen_font: ImageFont.ImageFont,
    width: int,
) -> list[str]:
    approximate_chars = max(6, int(width / max(1, chosen_font.size * 0.56)))
    return textwrap.wrap(value, width=approximate_chars, break_long_words=False) or [""]


def draw_cell_text(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    value: str,
    chosen_font: ImageFont.ImageFont,
    fill: str,
    align: str,
) -> None:
    x0, y0, x1, y1 = box
    lines = wrapped_lines(draw, value, chosen_font, x1 - x0 - 18)
    line_height = chosen_font.size + 4
    total_height = len(lines) * line_height
    y = y0 + max(5, (y1 - y0 - total_height) // 2)
    for line in lines:
        left, top, right, bottom = draw.textbbox((0, 0), line, font=chosen_font)
        text_width = right - left
        x = x0 + 9 if align == "left" else x1 - text_width - 9
        draw.text((x, y), line, font=chosen_font, fill=fill)
        y += line_height


def write_appendix_preview(table_data: pd.DataFrame) -> None:
    APPENDIX_PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    column_widths = [470, 570, 220, 570, 650, 500, 480, 520, 260, 320]
    width = sum(column_widths)
    title_height = 135
    header_height = 145
    row_height = 58
    height = title_height + header_height + row_height * len(table_data)
    image = Image.new("RGB", (width, height), f"#{WHITE}")
    draw = ImageDraw.Draw(image)
    title_font = font(42, bold=True)
    header_font = font(25, bold=True)
    body_font = font(23, bold=False)
    body_bold = font(23, bold=True)

    draw.rectangle((0, 0, width, title_height), fill=f"#{NAVY}")
    draw.text(
        (28, 39), APPENDIX_TITLE, font=title_font, fill=f"#{WHITE}", anchor=None
    )
    x = 0
    for column, header in enumerate(PREVIEW_HEADERS):
        x1 = x + column_widths[column]
        draw.rectangle((x, title_height, x1, title_height + header_height), fill=f"#{TEAL}")
        draw_cell_text(
            draw,
            (x, title_height, x1, title_height + header_height),
            header.replace("\n", " "),
            header_font,
            f"#{WHITE}",
            "left",
        )
        draw.line((x1, title_height, x1, height), fill=f"#{LIGHT_GREY}", width=2)
        x = x1

    for row_number, (_, record) in enumerate(table_data.iterrows()):
        y0 = title_height + header_height + row_number * row_height
        y1 = y0 + row_height
        background = PALE_BLUE if row_number % 2 == 0 else WHITE
        draw.rectangle((0, y0, width, y1), fill=f"#{background}")
        x = 0
        for column, header in enumerate(HEADERS):
            x1 = x + column_widths[column]
            is_left = column in (0, 1, 3, 4, 9)
            draw_cell_text(
                draw,
                (x, y0, x1, y1),
                format_preview_value(header, record[header]),
                body_bold if column in (0, 3) else body_font,
                f"#{NAVY if column in (0, 3) else TEXT}",
                "left" if is_left else "right",
            )
            draw.line((x1, y0, x1, y1), fill=f"#{LIGHT_GREY}", width=1)
            x = x1
        draw.line((0, y1, width, y1), fill=f"#{LIGHT_GREY}", width=1)
    image.save(APPENDIX_PREVIEW, dpi=(320, 320), optimize=True)


def write_main_preview(table_data: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    column_widths = [430, 560, 220, 490, 510, 500, 440, 480, 560]
    width = sum(column_widths)
    title_height = 135
    header_height = 145
    row_height = 66
    height = title_height + header_height + row_height * len(table_data)
    image = Image.new("RGB", (width, height), f"#{WHITE}")
    draw = ImageDraw.Draw(image)
    title_font = font(42, bold=True)
    header_font = font(24, bold=True)
    body_font = font(23, bold=False)
    body_bold = font(23, bold=True)

    draw.rectangle((0, 0, width, title_height), fill=f"#{NAVY}")
    draw.text((28, 39), TITLE, font=title_font, fill=f"#{WHITE}")
    x = 0
    for column, header in enumerate(MAIN_PREVIEW_HEADERS):
        x1 = x + column_widths[column]
        draw.rectangle((x, title_height, x1, title_height + header_height), fill=f"#{TEAL}")
        draw_cell_text(
            draw,
            (x, title_height, x1, title_height + header_height),
            header.replace("\n", " "),
            header_font,
            f"#{WHITE}",
            "left",
        )
        draw.line((x1, title_height, x1, height), fill=f"#{LIGHT_GREY}", width=2)
        x = x1

    for row_number, (_, record) in enumerate(table_data.iterrows()):
        y0 = title_height + header_height + row_number * row_height
        y1 = y0 + row_height
        background = PALE_BLUE if row_number % 2 == 0 else WHITE
        draw.rectangle((0, y0, width, y1), fill=f"#{background}")
        x = 0
        for column, header in enumerate(MAIN_HEADERS):
            x1 = x + column_widths[column]
            is_left = column in (0, 1)
            draw_cell_text(
                draw,
                (x, y0, x1, y1),
                format_preview_value(header, record[header]),
                body_bold if column == 0 else body_font,
                f"#{NAVY if column == 0 else TEXT}",
                "left" if is_left else "right",
            )
            draw.line((x1, y0, x1, y1), fill=f"#{LIGHT_GREY}", width=1)
            x = x1
        draw.line((0, y1, width, y1), fill=f"#{LIGHT_GREY}", width=1)
    image.save(PREVIEW, dpi=(320, 320), optimize=True)


def verify_main_outputs(row_count: int) -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == row_count + 2 and sheet.max_column == len(MAIN_HEADERS)
    assert sheet["A1"].value == TITLE
    assert not sheet.merged_cells.ranges
    assert sheet["A2"].value == MAIN_HEADERS[0]
    assert sheet["I2"].value == MAIN_HEADERS[-1]
    japanese_cells: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and JAPANESE_PATTERN.search(cell.value):
                japanese_cells.append(cell.coordinate)
    assert not japanese_cells, japanese_cells
    with Image.open(PREVIEW) as preview:
        dpi = preview.info.get("dpi", (0, 0))
        assert min(dpi) >= 300


def verify_appendix_outputs(row_count: int) -> None:
    workbook = load_workbook(APPENDIX_OUTPUT, data_only=False)
    sheet = workbook[APPENDIX_SHEET_NAME]
    assert sheet.max_row == row_count + 2 and sheet.max_column == len(HEADERS)
    assert sheet["A1"].value == APPENDIX_TITLE
    assert not sheet.merged_cells.ranges
    assert sheet["A2"].value == HEADERS[0]
    assert sheet["J2"].value == HEADERS[-1]
    japanese_cells: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and JAPANESE_PATTERN.search(cell.value):
                japanese_cells.append(cell.coordinate)
    assert not japanese_cells, japanese_cells
    with Image.open(APPENDIX_PREVIEW) as preview:
        dpi = preview.info.get("dpi", (0, 0))
        assert min(dpi) >= 300


def main() -> None:
    full_table, audit = construct_table()
    main_table = summarize_for_main_text(full_table)
    validate_table(full_table)
    validate_main_table(main_table)
    write_main_workbook(main_table)
    write_main_preview(main_table)
    write_appendix_workbook(full_table, audit)
    write_appendix_preview(full_table)
    verify_main_outputs(len(main_table))
    verify_appendix_outputs(len(full_table))
    print(
        f"Saved main {len(main_table)} rows x {len(MAIN_HEADERS)} cols -> "
        f"{OUTPUT.relative_to(ROOT)}"
    )
    print(f"Saved main PNG preview -> {PREVIEW.relative_to(ROOT)}")
    print(
        f"Saved appendix {len(full_table)} rows x {len(HEADERS)} cols -> "
        f"{APPENDIX_OUTPUT.relative_to(ROOT)}"
    )
    print(f"Saved appendix PNG preview -> {APPENDIX_PREVIEW.relative_to(ROOT)}")
    print(
        "Highest-load failed points by road state: "
        + "; ".join(
            f"{ROAD_SCENARIOS[key]}={value}"
            for key, value in audit["failure_names"].items()
        )
    )


if __name__ == "__main__":
    main()
