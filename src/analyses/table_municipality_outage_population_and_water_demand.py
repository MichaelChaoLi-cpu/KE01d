#!/usr/bin/env python3
"""Generate Municipality Outage Population and Water Demand.

Plan: Report municipalities with a positive outage in the selected official snapshot.
Framework: AnaSOP Sections 5-7 bounded affected-population and daily resident-
demand equations. The three demand columns hold the proportional central
population fixed and vary per-capita demand at 3, 10, and 20 L/person/day.
Zero, assumed-zero, and unmatched records remain in the processed audit data but
are not repeated in this main results table. Shelter demand remains separate.
"""

from pathlib import Path
import textwrap

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "data/processed/municipality_outage_demand_scenarios_preprocessed.parquet"
OUTPUT = ROOT / "data/results/tables/Table_municipality_outage_population_and_water_demand.xlsx"
PREVIEW = ROOT / "data/exp/Table_municipality_outage_population_and_water_demand.png"

TITLE = "Municipality Outage Population and Water Demand"
SHEET_NAME = "Municipality Demand"
HEADERS = [
    "Municipality",
    "Reported Outage Households",
    "Outage Household Ratio",
    "Lower Outage Population",
    "Central Outage Population",
    "Upper Outage Population",
    "Minimum Demand at Central Population (L/day)",
    "Basic Demand at Central Population (L/day)",
    "Extended Demand at Central Population (L/day)",
]

MUNICIPALITY_NAMES = {
    "43100": "Kumamoto City",
    "43202": "Yatsushiro City",
    "43203": "Hitoyoshi City",
    "43204": "Arao City",
    "43205": "Minamata City",
    "43206": "Tamana City",
    "43208": "Yamaga City",
    "43210": "Kikuchi City",
    "43211": "Uto City",
    "43212": "Kamiamakusa City",
    "43213": "Uki City",
    "43214": "Aso City",
    "43215": "Amakusa City",
    "43216": "Koshi City",
    "43348": "Misato Town",
    "43364": "Gyokuto Town",
    "43367": "Nankan Town",
    "43368": "Nagasu Town",
    "43369": "Nagomi Town",
    "43403": "Ozu Town",
    "43404": "Kikuyo Town",
    "43423": "Minamioguni Town",
    "43424": "Oguni Town",
    "43425": "Ubuyama Village",
    "43428": "Takamori Town",
    "43432": "Nishihara Village",
    "43433": "Minamiaso Village",
    "43441": "Mifune Town",
    "43442": "Kashima Town",
    "43443": "Mashiki Town",
    "43444": "Kosa Town",
    "43447": "Yamato Town",
    "43468": "Hikawa Town",
    "43482": "Ashikita Town",
    "43484": "Tsunagi Town",
    "43501": "Nishiki Town",
    "43505": "Taragi Town",
    "43506": "Yunomae Town",
    "43507": "Mizukami Village",
    "43510": "Sagara Village",
    "43511": "Itsuki Village",
    "43512": "Yamae Village",
    "43513": "Kuma Village",
    "43514": "Asagiri Town",
    "43531": "Reihoku Town",
}

OUTAGE_STATUS_LABELS = {
    "reported_positive": "Reported positive outage",
    "reported_zero": "Reported zero outage",
    "assumed_zero_no_official_outage_listing": "Assumed zero: absent from official listing",
    "unmatched_geography": "Unmatched geography",
}
NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"


def read_source() -> pd.DataFrame:
    columns = [
        "Reporting Municipality Code",
        "Report Number",
        "Water Status Timestamp",
        "Current Outage Households",
        "Outage Observation Status",
        "Outage Household Ratio",
        "Outage Population Scenario",
        "Estimated Outage Population",
        "Demand Scenario",
        "Per Capita Water Demand (L/person/day)",
        "Estimated Water Demand (L/day)",
    ]
    data = pd.read_parquet(INPUT, columns=columns)
    assert data.shape[0] == 414, f"Expected 414 long-format rows, found {len(data)}"
    assert data["Reporting Municipality Code"].dropna().nunique() == 45
    assert set(data["Outage Observation Status"]) == set(OUTAGE_STATUS_LABELS)
    return data


def first_value(frame: pd.DataFrame, column: str):
    values = frame[column].drop_duplicates()
    assert len(values) == 1, f"Non-unique {column}: {values.tolist()}"
    return values.iloc[0]


def construct_table(source: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    source = source.copy()
    source["_unit"] = source["Reporting Municipality Code"].fillna("AUDIT-UNMATCHED")

    population_scenarios = {
        "lower_one_person_per_household": "Lower Outage Population",
        "proportional_central": "Central Outage Population",
        "upper_p90_household_size": "Upper Outage Population",
    }
    demand_scenarios = {
        "minimum": "Minimum Demand at Central Population (L/day)",
        "basic": "Basic Demand at Central Population (L/day)",
        "extended": "Extended Demand at Central Population (L/day)",
    }

    for unit, frame in source.groupby("_unit", sort=False, dropna=False):
        code = str(unit)
        unmatched = code == "AUDIT-UNMATCHED"
        status_code = first_value(frame, "Outage Observation Status")
        first_value(frame, "Water Status Timestamp")
        row: dict[str, object] = {
            "Municipality": "Unmatched Mesh Audit Unit"
            if unmatched
            else MUNICIPALITY_NAMES[code],
            "Reported Outage Households": first_value(
                frame, "Current Outage Households"
            ),
            "Outage Household Ratio": first_value(frame, "Outage Household Ratio"),
        }

        for scenario, column in population_scenarios.items():
            values = frame.loc[
                frame["Outage Population Scenario"].eq(scenario),
                "Estimated Outage Population",
            ].drop_duplicates()
            assert len(values) == 1, (unit, scenario, values.tolist())
            row[column] = values.iloc[0]

        central = frame.loc[
            frame["Outage Population Scenario"].eq("proportional_central")
        ]
        for scenario, column in demand_scenarios.items():
            scenario_frame = central.loc[central["Demand Scenario"].eq(scenario)]
            values = scenario_frame["Estimated Water Demand (L/day)"].drop_duplicates()
            assert len(values) == 1, (unit, scenario, values.tolist())
            row[column] = values.iloc[0]

        if status_code == "reported_positive":
            rows.append(row)

    table = pd.DataFrame(rows, columns=HEADERS)
    table = (
        table.sort_values(
            ["Reported Outage Households", "Municipality"],
            ascending=[False, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )
    return table


def validate_table(table: pd.DataFrame) -> None:
    assert table.shape == (3, 9), table.shape
    assert table.columns.tolist() == HEADERS
    assert table["Municipality"].tolist() == [
        "Yatsushiro City",
        "Uki City",
        "Hikawa Town",
    ]
    assert table["Reported Outage Households"].gt(0).all()

    known = table["Central Outage Population"].notna()
    assert (
        table.loc[known, "Lower Outage Population"]
        <= table.loc[known, "Central Outage Population"]
    ).all()
    assert (
        table.loc[known, "Central Outage Population"]
        <= table.loc[known, "Upper Outage Population"]
    ).all()
    for column, liters in [
        ("Minimum Demand at Central Population (L/day)", 3),
        ("Basic Demand at Central Population (L/day)", 10),
        ("Extended Demand at Central Population (L/day)", 20),
    ]:
        expected = table.loc[known, "Central Outage Population"] * liters
        observed = table.loc[known, column]
        assert (observed - expected).abs().max() < 1e-6

def write_workbook(table_data: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False

    # Preserve the title in row 1 without merged cells for deterministic DOCX export.
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 10):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 58

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(table_data.iterrows()):
        excel_row = 3 + offset
        fill = PALE_BLUE if offset % 2 == 0 else WHITE
        for column, header in enumerate(HEADERS, start=1):
            value = record[header]
            if pd.isna(value):
                value = None
            cell = sheet.cell(excel_row, column, value)
            cell.font = Font(
                name="Aptos",
                size=9.3,
                bold=column == 1,
                color=NAVY if column == 1 else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "right",
                vertical="center",
                wrap_text=column == 1,
            )
        for column in (2, 4, 5, 6, 7, 8, 9):
            sheet.cell(excel_row, column).number_format = "#,##0"
        sheet.cell(excel_row, 3).number_format = "0.0%"
        sheet.row_dimensions[excel_row].height = 38

    end_row = len(table_data) + 2
    excel_table = Table(displayName="MunicipalityOutageDemand", ref=f"A2:I{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)

    widths = {
        "A": 25,
        "B": 24,
        "C": 21,
        "D": 23,
        "E": 24,
        "F": 23,
        "G": 32,
        "H": 32,
        "I": 32,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "B3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:I{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · Resident demand only · Central population used for 3/10/20 L demand columns"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY

    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d municipality outage-population and resident-demand scenarios"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = "Kumamoto, emergency water, outage population, municipal demand"
    workbook.save(OUTPUT)


def preview_rows(table_data: pd.DataFrame) -> list[list[str]]:
    wrap_widths = [20, 18, 16, 18, 19, 18, 24, 24, 24]
    integer_columns = {
        "Reported Outage Households",
        "Lower Outage Population",
        "Central Outage Population",
        "Upper Outage Population",
        "Minimum Demand at Central Population (L/day)",
        "Basic Demand at Central Population (L/day)",
        "Extended Demand at Central Population (L/day)",
    }
    rows: list[list[str]] = []
    for _, record in table_data.iterrows():
        row: list[str] = []
        for header, width in zip(HEADERS, wrap_widths, strict=True):
            value = record[header]
            if pd.isna(value) or value == "":
                text = "—"
            elif header in integer_columns:
                text = f"{float(value):,.0f}"
            elif header == "Outage Household Ratio":
                text = f"{float(value):.1%}"
            else:
                text = str(value)
            row.append(textwrap.fill(text, width=width, break_long_words=False))
        rows.append(row)
    return rows


def write_preview(table_data: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 7})
    figure, axis = plt.subplots(figsize=(22, 5.2))
    axis.axis("off")

    axis.add_patch(
        Rectangle(
            (0, 0.84), 1, 0.16,
            transform=axis.transAxes,
            facecolor=f"#{NAVY}", edgecolor=f"#{NAVY}", clip_on=False,
        )
    )
    axis.text(
        0.010, 0.92, TITLE,
        transform=axis.transAxes,
        ha="left", va="center",
        fontsize=18, fontweight="bold", color="white",
    )

    column_widths = [0.12, 0.11, 0.10, 0.105, 0.11, 0.105, 0.15, 0.15, 0.15]
    preview = axis.table(
        cellText=preview_rows(table_data),
        colLabels=[textwrap.fill(header, 22) for header in HEADERS],
        colWidths=column_widths,
        cellLoc="left",
        colLoc="center",
        bbox=[0, 0, 1, 0.84],
    )
    preview.auto_set_font_size(False)
    preview.set_fontsize(8.2)

    header_height = 0.22
    body_height = (0.84 - header_height) / len(table_data)
    for (row, column), cell in preview.get_celld().items():
        cell.set_edgecolor(f"#{LIGHT_GREY}")
        cell.set_linewidth(0.45)
        if row == 0:
            cell.set_facecolor(f"#{TEAL}")
            cell.set_text_props(color="white", weight="bold", ha="center", va="center")
            cell.set_height(header_height)
        else:
            cell.set_facecolor(f"#{PALE_BLUE if (row - 1) % 2 == 0 else WHITE}")
            numeric = column in (1, 2, 3, 4, 5, 6, 7, 8)
            cell.set_text_props(
                color=f"#{NAVY}" if column == 0 else f"#{TEXT}",
                weight="bold" if column == 0 else "normal",
                ha="right" if numeric else "left",
                va="center",
            )
            cell.set_height(body_height)
        cell.PAD = 0.025

    figure.savefig(PREVIEW, dpi=320, bbox_inches="tight", pad_inches=0.06)
    plt.close(figure)


def verify_outputs() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == 5
    assert sheet.max_column == 9
    assert sheet["A1"].value == TITLE
    assert not sheet.merged_cells.ranges
    assert sheet["A2"].value == HEADERS[0]
    assert sheet["I2"].value == HEADERS[-1]
    assert sum(1 for row in sheet.iter_rows(min_row=3, max_row=5) if row[0].value) == 3
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("#")
        for row in sheet.iter_rows()
        for cell in row
    )
    assert PREVIEW.exists() and PREVIEW.stat().st_size > 0


def main() -> None:
    table_data = construct_table(read_source())
    validate_table(table_data)
    write_workbook(table_data)
    write_preview(table_data)
    verify_outputs()
    known = table_data["Central Outage Population"].notna()
    print(f"Saved {len(table_data)} rows x {len(HEADERS)} cols -> {OUTPUT.relative_to(ROOT)}")
    print(f"Saved faithful PNG preview -> {PREVIEW.relative_to(ROOT)}")
    print(
        "Central affected population: "
        f"{table_data.loc[known, 'Central Outage Population'].sum():,.0f}; "
        "minimum resident demand: "
        f"{table_data.loc[known, 'Minimum Demand at Central Population (L/day)'].sum():,.0f} L/day"
    )


if __name__ == "__main__":
    main()
