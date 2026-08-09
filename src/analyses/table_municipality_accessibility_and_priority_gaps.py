#!/usr/bin/env python3
"""Generate Municipality Accessibility and Priority Gaps.

Plan: Compare nominal road-network access for the three municipalities with a
reported positive outage. Framework: AnaSOP Sections 5-7 shortest-network-
distance and weighted-coverage equations. Resident coverage is reported at the
co-equal 250, 500, and 1,000 m thresholds; age-65-plus priority coverage uses
the strict 250 m threshold; shelter coverage uses 1,000 m. Network-undefined
units remain in denominators and are therefore counted as uncovered. The
publication table is transposed so indicators form rows and municipalities form
columns.
"""

from pathlib import Path
import textwrap

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from figure_accessibility_coverage_by_distance_threshold import attach_unit_distances
from figure_announced_water_points_and_nominal_access_coverage import (
    build_baseline_graph,
    nearest_water_point_node_distances,
)
from table_municipality_outage_population_and_water_demand import MUNICIPALITY_NAMES


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
OUTPUT = ROOT / "data/results/tables/Table_municipality_accessibility_and_priority_gaps.xlsx"
PREVIEW = ROOT / "data/exp/Table_municipality_accessibility_and_priority_gaps.png"

TITLE = "Municipality Accessibility and Priority Gaps"
SHEET_NAME = "Municipality Access"
DISTANCE = "Nearest Water Point Network Distance (m)"
POSITIVE_CODES = ["43202", "43213", "43468"]
HEADERS = [
    "Municipality",
    "Affected Residents",
    "Resident Coverage <=250 m",
    "Resident Coverage <=500 m",
    "Resident Coverage <=1,000 m",
    "Residents Uncovered at 1,000 m",
    "Residents Age 65+ in Outage Municipality",
    "Age 65+ Coverage <=250 m",
    "Age 65+ Uncovered at 250 m",
    "Shelter Evacuees",
    "Shelter Coverage <=1,000 m",
    "Shelter Evacuees Uncovered at 1,000 m",
]
DISPLAY_HEADERS = ["Indicator", "Yatsushiro City", "Uki City", "Hikawa Town"]

NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"


def read(name: str, columns: list[str] | None = None) -> pd.DataFrame:
    return pd.read_parquet(PROCESSED / name, columns=columns)


def build_distances() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    water = read("emergency_water_points_network_access_preprocessed.parquet")
    eligible_water = water.loc[
        water["Network Snap Accepted"].fillna(False)
        & water["Water Point Node ID"].notna()
    ].copy()
    assert len(water) == 36 and len(eligible_water) == 36

    nodes = read("kumamoto_routable_road_nodes_preprocessed.parquet", ["Network Node ID"])
    edges = read(
        "kumamoto_routable_road_edges_preprocessed.parquet",
        [
            "Road Edge ID",
            "From Node ID",
            "To Node ID",
            "Road Length (m)",
            "Road Available",
            "Network Analysis Eligible",
        ],
    )
    graph, node_index, eligible_edges = build_baseline_graph(nodes, edges)
    node_distances = nearest_water_point_node_distances(
        graph, node_index, eligible_water
    )

    mesh_scenarios = read(
        "population_mesh_outage_demand_scenarios_preprocessed.parquet",
        [
            "Mesh Code",
            "Reporting Municipality Code",
            "Outage Population Scenario",
            "Demand Scenario",
            "Outage Household Ratio",
            "Estimated Outage Population",
        ],
    )
    mesh_scenarios = mesh_scenarios.loc[
        mesh_scenarios["Outage Population Scenario"].eq("proportional_central")
        & mesh_scenarios["Demand Scenario"].eq("minimum")
    ].copy()

    mesh_access = read("kumamoto_population_mesh_network_access_preprocessed.parquet").drop(
        columns=["Geometry"], errors="ignore"
    )
    residents = mesh_scenarios.merge(
        mesh_access, on="Mesh Code", how="left", validate="one_to_one"
    )
    residents = residents.loc[
        residents["Reporting Municipality Code"].isin(POSITIVE_CODES)
        & residents["Estimated Outage Population"].gt(0)
    ].copy()
    residents = attach_unit_distances(residents, eligible_edges, node_distances)

    groups = read(
        "kumamoto_population_disclosure_groups_preprocessed.parquet",
        ["Disclosure Group Code", "Population Age 65+"],
    )
    group_access = read(
        "kumamoto_population_group_network_access_preprocessed.parquet"
    ).drop(columns=["Geometry"], errors="ignore")
    older = groups.merge(
        group_access, on="Disclosure Group Code", how="inner", validate="one_to_one"
    )
    older = older.merge(
        mesh_scenarios[
            ["Mesh Code", "Reporting Municipality Code", "Outage Household Ratio"]
        ],
        left_on="Representative Mesh Code",
        right_on="Mesh Code",
        how="left",
        validate="many_to_one",
    )
    older = older.loc[
        older["Reporting Municipality Code"].isin(POSITIVE_CODES)
        & older["Outage Household Ratio"].gt(0)
        & older["Population Age 65+"].gt(0)
    ].copy()
    older = attach_unit_distances(older, eligible_edges, node_distances)

    shelters = read("shelter_water_demand_network_access_preprocessed.parquet")
    shelters = shelters.loc[shelters["Demand Scenario"].eq("minimum")].drop_duplicates(
        "Shelter Number"
    )
    shelter_code = {"八代市": "43202", "宇城市": "43213", "氷川町": "43468"}
    shelters["Reporting Municipality Code"] = shelters["Municipality"].map(shelter_code)
    shelters = shelters.loc[
        shelters["Reporting Municipality Code"].isin(POSITIVE_CODES)
    ].copy()
    shelters = attach_unit_distances(shelters, eligible_edges, node_distances)
    return residents, older, shelters


def weighted_summary(
    frame: pd.DataFrame,
    weight_column: str,
    thresholds: list[float],
) -> tuple[float, dict[float, float], dict[float, float]]:
    weight = pd.to_numeric(frame[weight_column], errors="coerce").fillna(0)
    distance = pd.to_numeric(frame[DISTANCE], errors="coerce")
    denominator = float(weight.sum())
    if denominator <= 0:
        return 0.0, {threshold: np.nan for threshold in thresholds}, {
            threshold: 0.0 for threshold in thresholds
        }
    covered = {
        threshold: float(weight[distance.notna() & distance.le(threshold)].sum())
        for threshold in thresholds
    }
    coverage = {
        threshold: covered[threshold] / denominator for threshold in thresholds
    }
    uncovered = {
        threshold: denominator - covered[threshold] for threshold in thresholds
    }
    return denominator, coverage, uncovered


def construct_table() -> pd.DataFrame:
    residents, older, shelters = build_distances()
    rows: list[dict[str, object]] = []
    for code in POSITIVE_CODES:
        resident_unit = residents.loc[residents["Reporting Municipality Code"].eq(code)]
        older_unit = older.loc[older["Reporting Municipality Code"].eq(code)]
        shelter_unit = shelters.loc[shelters["Reporting Municipality Code"].eq(code)]

        resident_total, resident_coverage, resident_uncovered = weighted_summary(
            resident_unit, "Estimated Outage Population", [250, 500, 1000]
        )
        older_total, older_coverage, older_uncovered = weighted_summary(
            older_unit, "Population Age 65+", [250]
        )
        shelter_total, shelter_coverage, shelter_uncovered = weighted_summary(
            shelter_unit, "Evacuee People", [1000]
        )
        rows.append(
            {
                "Municipality": MUNICIPALITY_NAMES[code],
                "Affected Residents": resident_total,
                "Resident Coverage <=250 m": resident_coverage[250],
                "Resident Coverage <=500 m": resident_coverage[500],
                "Resident Coverage <=1,000 m": resident_coverage[1000],
                "Residents Uncovered at 1,000 m": resident_uncovered[1000],
                "Residents Age 65+ in Outage Municipality": older_total,
                "Age 65+ Coverage <=250 m": older_coverage[250],
                "Age 65+ Uncovered at 250 m": older_uncovered[250],
                "Shelter Evacuees": shelter_total,
                "Shelter Coverage <=1,000 m": shelter_coverage[1000],
                "Shelter Evacuees Uncovered at 1,000 m": shelter_uncovered[1000],
            }
        )
    return pd.DataFrame(rows, columns=HEADERS)


def validate_table(table_data: pd.DataFrame) -> None:
    assert table_data.shape == (3, 12), table_data.shape
    assert table_data["Municipality"].tolist() == [
        "Yatsushiro City", "Uki City", "Hikawa Town"
    ]
    coverage_columns = [column for column in HEADERS if "Coverage" in column]
    for column in coverage_columns:
        values = table_data[column].dropna()
        assert values.between(0, 1).all(), column
    assert (
        table_data["Resident Coverage <=250 m"]
        <= table_data["Resident Coverage <=500 m"]
    ).all()
    assert (
        table_data["Resident Coverage <=500 m"]
        <= table_data["Resident Coverage <=1,000 m"]
    ).all()
    assert table_data["Affected Residents"].sum() > 93_000
    assert table_data["Shelter Evacuees"].sum() == 2_248
    assert table_data.loc[
        table_data["Municipality"].ne("Yatsushiro City"),
        "Shelter Coverage <=1,000 m",
    ].isna().all()


def transpose_table(table_data: pd.DataFrame) -> pd.DataFrame:
    displayed = (
        table_data.set_index("Municipality")
        .transpose()
        .reset_index()
        .rename(columns={"index": "Indicator"})
    )
    displayed = displayed[DISPLAY_HEADERS]
    assert displayed.shape == (11, 4)
    assert displayed["Indicator"].tolist() == HEADERS[1:]
    return displayed


def group_fill(indicator: str) -> str:
    if "Age 65+" in indicator:
        return WHITE
    if indicator.startswith("Resident") or indicator == "Affected Residents":
        return PALE_BLUE
    return PALE_BLUE


def write_workbook(displayed: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:D1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 5):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(DISPLAY_HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 38

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(displayed.iterrows()):
        excel_row = 3 + offset
        indicator = record["Indicator"]
        fill = group_fill(indicator)
        for column, header in enumerate(DISPLAY_HEADERS, start=1):
            value = record[header]
            if pd.isna(value):
                value = None
            cell = sheet.cell(excel_row, column, value)
            cell.font = Font(
                name="Aptos", size=9.3,
                bold=column == 1,
                color=NAVY if column == 1 else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "right",
                vertical="center", wrap_text=column == 1,
            )
        number_format = "0.0%" if "Coverage" in indicator else "#,##0"
        for column in (2, 3, 4):
            sheet.cell(excel_row, column).number_format = number_format
        sheet.row_dimensions[excel_row].height = 34

    end_row = len(displayed) + 2
    excel_table = Table(displayName="MunicipalityAccessibilityGaps", ref=f"A2:D{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)
    coverage_excel_rows = [
        3 + index
        for index, indicator in enumerate(displayed["Indicator"])
        if "Coverage" in indicator
    ]
    for excel_row in coverage_excel_rows:
        sheet.conditional_formatting.add(
            f"B{excel_row}:D{excel_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F4CCCC",
                mid_type="num", mid_value=0.5, mid_color="FFF2CC",
                end_type="num", end_value=1, end_color="D9EAD3",
            ),
        )

    widths = {"A": 48, "B": 24, "C": 24, "D": 24}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:D{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · Baseline road-network proxy · Undefined distances count as uncovered"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY
    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d municipality nominal-access and priority gaps"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = "Kumamoto, emergency water, accessibility, priority population"
    workbook.save(OUTPUT)


def preview_rows(displayed: pd.DataFrame) -> list[list[str]]:
    rows: list[list[str]] = []
    for _, record in displayed.iterrows():
        indicator = record["Indicator"]
        row = [indicator]
        for municipality in DISPLAY_HEADERS[1:]:
            value = record[municipality]
            if pd.isna(value):
                text = "—"
            elif "Coverage" in indicator:
                text = f"{float(value):.1%}"
            else:
                text = f"{float(value):,.0f}"
            row.append(text)
        rows.append(row)
    return rows


def write_preview(displayed: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 7})
    figure, axis = plt.subplots(figsize=(15, 10.5))
    axis.axis("off")
    axis.add_patch(
        Rectangle(
            (0, 0.91), 1, 0.09,
            transform=axis.transAxes,
            facecolor=f"#{NAVY}", edgecolor=f"#{NAVY}", clip_on=False,
        )
    )
    axis.text(
        0.018, 0.955, TITLE,
        transform=axis.transAxes,
        ha="left", va="center",
        fontsize=18, fontweight="bold", color="white",
    )
    column_widths = [0.40, 0.20, 0.20, 0.20]
    preview = axis.table(
        cellText=preview_rows(displayed),
        colLabels=DISPLAY_HEADERS,
        colWidths=column_widths,
        cellLoc="left", colLoc="center", bbox=[0, 0, 1, 0.91],
    )
    preview.auto_set_font_size(False)
    preview.set_fontsize(9.2)
    header_height = 0.085
    body_height = (0.91 - header_height) / len(displayed)
    for (row, column), cell in preview.get_celld().items():
        cell.set_edgecolor(f"#{LIGHT_GREY}")
        cell.set_linewidth(0.45)
        if row == 0:
            cell.set_facecolor(f"#{TEAL}")
            cell.set_text_props(color="white", weight="bold", ha="center", va="center")
            cell.set_height(header_height)
        else:
            indicator = displayed.iloc[row - 1]["Indicator"]
            cell.set_facecolor(f"#{group_fill(indicator)}")
            cell.set_text_props(
                color=f"#{NAVY}" if column == 0 else f"#{TEXT}",
                weight="bold" if column == 0 else "normal",
                ha="left" if column == 0 else "right", va="center",
            )
            cell.set_height(body_height)
        cell.PAD = 0.025
    figure.savefig(PREVIEW, dpi=320, bbox_inches="tight", pad_inches=0.06)
    plt.close(figure)


def verify_outputs() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == 13 and sheet.max_column == 4
    assert sheet["A1"].value == TITLE
    assert sheet["A2"].value == DISPLAY_HEADERS[0]
    assert sheet["D2"].value == DISPLAY_HEADERS[-1]
    assert sum(1 for row in sheet.iter_rows(min_row=3, max_row=13) if row[0].value) == 11
    assert PREVIEW.exists() and PREVIEW.stat().st_size > 0


def main() -> None:
    table_data = construct_table()
    validate_table(table_data)
    displayed = transpose_table(table_data)
    write_workbook(displayed)
    write_preview(displayed)
    verify_outputs()
    print(f"Saved {len(displayed)} rows x {len(DISPLAY_HEADERS)} cols -> {OUTPUT.relative_to(ROOT)}")
    print(f"Saved faithful PNG preview -> {PREVIEW.relative_to(ROOT)}")
    print(table_data.to_string(index=False, float_format=lambda value: f"{value:.4f}"))


if __name__ == "__main__":
    main()
