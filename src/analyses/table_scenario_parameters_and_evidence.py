#!/usr/bin/env python3
"""Generate the planned Scenario Parameters and Evidence table.

The Excel workbook is the authoritative research table. The PNG is a faithful
preview of the same 37 rows and seven columns; it does not remove, narrow, or
split content for mobile viewing.
"""

from pathlib import Path
import textwrap

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "data/processed/emergency_water_scenario_parameters_preprocessed.parquet"
OUTPUT = ROOT / "data/results/tables/Table_scenario_parameters_and_evidence.xlsx"
PREVIEW = ROOT / "data/exp/Table_scenario_parameters_and_evidence.png"

TITLE = "Scenario Parameters and Evidence"
SHEET_NAME = "Scenario Parameters"
HEADERS = [
    "Parameter Name",
    "Scenario Level",
    "Parameter Value",
    "Parameter Unit",
    "Evidence Class",
    "Evidence Source",
    "Parameter Notes",
]

EVIDENCE_CLASS_LABELS = {
    "mixed_official_and_research_sensitivity": "Mixed official and research sensitivity",
    "official_reference_sensitivity": "Official reference sensitivity",
    "researcher_defined_sensitivity": "Researcher-defined sensitivity",
}

NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
PALE_TEAL = "E8F3F3"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"


def read_and_validate() -> pd.DataFrame:
    data = pd.read_parquet(INPUT)
    assert data.columns.tolist() == HEADERS, data.columns.tolist()
    assert data.shape == (37, 7), f"Expected 37 rows x 7 columns, found {data.shape}"
    assert not data[HEADERS].isna().any().any(), "Missing table value"
    assert not data.duplicated(["Parameter Name", "Scenario Level"]).any(), (
        "Duplicate parameter/scenario rows"
    )

    nominal = data[data["Parameter Name"].eq("General Access Distance")]
    assert nominal["Scenario Level"].tolist() == [
        "250 m scenario",
        "500 m scenario",
        "1,000 m scenario",
    ]
    assert nominal["Parameter Value"].astype(int).tolist() == [250, 500, 1000]
    assert not nominal["Scenario Level"].str.contains(
        "central|baseline|reference", case=False, regex=True
    ).any()

    extended = data[
        data["Parameter Name"].eq("Extended Access Distance Diagnostic")
    ]
    assert extended["Parameter Value"].astype(int).tolist() == [2000, 5000]
    assert extended["Scenario Level"].str.contains("diagnostic", case=False).all()
    return data


def group_fills(data: pd.DataFrame) -> list[str]:
    """Alternate fill by parameter block so scenario rows stay visually grouped."""
    fills: list[str] = []
    previous = None
    group_index = -1
    for parameter in data["Parameter Name"]:
        if parameter != previous:
            group_index += 1
            previous = parameter
        fills.append(PALE_BLUE if group_index % 2 == 0 else WHITE)
    return fills


def display_values(data: pd.DataFrame) -> pd.DataFrame:
    """Convert machine-readable evidence codes to publication-readable phrases."""
    displayed = data.copy()
    unknown = set(displayed["Evidence Class"]) - set(EVIDENCE_CLASS_LABELS)
    assert not unknown, f"Unmapped evidence classes: {sorted(unknown)}"
    displayed["Evidence Class"] = displayed["Evidence Class"].map(EVIDENCE_CLASS_LABELS)
    assert not displayed["Evidence Class"].str.contains("_", regex=False).any()
    return displayed


def write_workbook(data: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = TITLE
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 8):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 36

    thin = Side(style="thin", color=LIGHT_GREY)
    fills = group_fills(data)
    previous_parameter = None
    for offset, (_, record) in enumerate(data.iterrows()):
        excel_row = 3 + offset
        parameter = record["Parameter Name"]
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(excel_row, column, record[header])
            cell.font = Font(
                name="Aptos",
                size=9.5,
                bold=column == 1 and parameter != previous_parameter,
                color=NAVY if column == 1 else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fills[offset])
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column in (1, 2, 5, 6, 7) else "center",
                vertical="top",
                wrap_text=True,
            )
        source_length = len(str(record["Evidence Source"]))
        note_length = len(str(record["Parameter Notes"]))
        sheet.row_dimensions[excel_row].height = min(
            84, max(42, 42 + 10 * max(source_length // 85, note_length // 95))
        )
        previous_parameter = parameter

    end_row = len(data) + 2
    table = Table(displayName="ScenarioParametersEvidence", ref=f"A2:G{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    widths = {
        "A": 34,
        "B": 28,
        "C": 20,
        "D": 20,
        "E": 42,
        "F": 76,
        "G": 82,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:G{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = "KE01d · Research planning scenarios · Generated from processed data"
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY

    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d emergency-water planning parameters and evidence"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = "Kumamoto, emergency water, scenario parameters, evidence"
    workbook.save(OUTPUT)


def wrapped_preview_data(data: pd.DataFrame) -> list[list[str]]:
    widths = {
        "Parameter Name": 28,
        "Scenario Level": 24,
        "Parameter Value": 18,
        "Parameter Unit": 17,
        "Evidence Class": 30,
        "Evidence Source": 49,
        "Parameter Notes": 54,
    }
    rows: list[list[str]] = []
    for _, record in data.iterrows():
        rows.append(
            [
                textwrap.fill(str(record[header]), width=widths[header], break_long_words=True)
                for header in HEADERS
            ]
        )
    return rows


def write_preview(data: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 8})
    figure, axis = plt.subplots(figsize=(22, 27))
    axis.axis("off")

    axis.add_patch(
        Rectangle(
            (0, 0.966),
            1,
            0.034,
            transform=axis.transAxes,
            facecolor=f"#{NAVY}",
            edgecolor=f"#{NAVY}",
            clip_on=False,
        )
    )
    axis.text(
        0.012,
        0.983,
        TITLE,
        transform=axis.transAxes,
        ha="left",
        va="center",
        fontsize=18,
        fontweight="bold",
        color="white",
    )

    column_widths = [0.15, 0.125, 0.085, 0.085, 0.16, 0.195, 0.20]
    table = axis.table(
        cellText=wrapped_preview_data(data),
        colLabels=[textwrap.fill(header, 22) for header in HEADERS],
        colWidths=column_widths,
        cellLoc="left",
        colLoc="center",
        bbox=[0, 0, 1, 0.966],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(7.4)

    header_height = 0.040
    body_height = (0.966 - header_height) / len(data)
    fills = group_fills(data)
    for (row, column), cell in table.get_celld().items():
        cell.set_edgecolor(f"#{LIGHT_GREY}")
        cell.set_linewidth(0.5)
        if row == 0:
            cell.set_facecolor(f"#{TEAL}")
            cell.set_text_props(color="white", weight="bold", ha="center", va="center")
            cell.set_height(header_height)
        else:
            cell.set_facecolor(f"#{fills[row - 1]}")
            cell.set_text_props(
                color=f"#{NAVY}" if column == 0 else f"#{TEXT}",
                weight="bold"
                if column == 0
                and (
                    row == 1
                    or data.iloc[row - 1]["Parameter Name"]
                    != data.iloc[row - 2]["Parameter Name"]
                )
                else "normal",
                ha="center" if column in (2, 3) else "left",
                va="center",
            )
            cell.set_height(body_height)
        cell.PAD = 0.026

    figure.savefig(PREVIEW, dpi=320, bbox_inches="tight", pad_inches=0.06)
    plt.close(figure)


def verify_workbook() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == 39
    assert sheet.max_column == 7
    assert sheet["A1"].value == TITLE
    assert sheet["A2"].value == HEADERS[0]
    assert sheet["G2"].value == HEADERS[-1]
    assert sum(1 for row in sheet.iter_rows(min_row=3, max_row=39) if row[0].value) == 37
    assert not any("_" in str(sheet.cell(row, 5).value) for row in range(3, 40))
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("#")
        for row in sheet.iter_rows()
        for cell in row
    )
    assert PREVIEW.exists() and PREVIEW.stat().st_size > 0


def main() -> None:
    data = display_values(read_and_validate())
    write_workbook(data)
    write_preview(data)
    verify_workbook()
    print(f"Saved {len(data)} rows x {len(HEADERS)} cols -> {OUTPUT.relative_to(ROOT)}")
    print(f"Saved faithful PNG preview -> {PREVIEW.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
