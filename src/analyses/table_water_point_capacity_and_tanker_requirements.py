#!/usr/bin/env python3
"""Generate Water-Point Capacity and Tanker Requirements.

Plan: Report announced water points receiving positive resident demand under a
transparent nearest-point assignment within the 2,000 m extended allocation
diagnostic. Framework: AnaSOP Sections 5-7 central affected population,
3 L/person/day minimum demand, best historical refill-route productivity,
announced operating windows, 3,000 L tanker capacity, five-trip cap, 10-hour
work limit, and 60 minutes total loading/unloading time. Results are required
planning capacity, not observed deliveries, available fleet, or point capacity.
"""

from pathlib import Path
import math
import textwrap

import geopandas as gpd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from figure_scenario_based_tanker_and_temporary_water_point_allocation import (
    build_graphs,
    demand_units,
    distance_arcs,
    route_productivity,
)
from table_municipality_outage_population_and_water_demand import MUNICIPALITY_NAMES
from table_scenario_based_priority_deployment_list import (
    ENGLISH_SITE_NAMES,
    JAPANESE_CHARACTER_PATTERN,
    translate_exact,
)


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
OUTPUT = ROOT / "data/results/tables/Table_water_point_capacity_and_tanker_requirements.xlsx"
PREVIEW = ROOT / "data/exp/Table_water_point_capacity_and_tanker_requirements.png"

TITLE = "Water-Point Capacity and Tanker Requirements"
SHEET_NAME = "Point Requirements"
ACCESS_DISTANCE_M = 2_000.0
HEADERS = [
    "Municipality",
    "Water Point",
    "Operating Window",
    "Assigned Affected Residents",
    "Required Water (L/day)",
    "Best-Route Cycle (min/trip)",
    "Required Daily Trips",
    "Tanker-Day Capacity Required (15,000 L/day)",
]
PREVIEW_HEADERS = [
    "Municipality",
    "Water Point",
    "Operating\nWindow",
    "Assigned Affected\nResidents",
    "Required Water\n(L/day)",
    "Best-Route Cycle\n(min/trip)",
    "Required\nDaily Trips",
    "Tanker-Day Capacity Required\n(15,000 L/day)",
]

MUNICIPALITY_CODE_BY_JAPANESE_NAME = {
    "八代市": "43202",
    "宇城市": "43213",
    "氷川町": "43468",
}

NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"


def read(name: str, columns: list[str] | None = None):
    path = PROCESSED / name
    if "Geometry" in (columns or []):
        return gpd.read_parquet(path, columns=columns)
    return pd.read_parquet(path, columns=columns)


def operating_minutes(opening: str, closing: str) -> float:
    open_hour, open_minute = (int(value) for value in str(opening).split(":"))
    close_hour, close_minute = (int(value) for value in str(closing).split(":"))
    duration = (close_hour * 60 + close_minute) - (open_hour * 60 + open_minute)
    assert duration > 0
    return float(duration)


def build_point_inventory(
    water: gpd.GeoDataFrame,
    node_index: pd.Series,
) -> gpd.GeoDataFrame:
    points = water.loc[
        water["Network Snap Accepted"].fillna(False)
        & water["Water Point Node ID"].notna()
    ].copy().reset_index(drop=True)
    assert len(points) == 36
    points["Site ID"] = "WATER::" + points.index.astype(str)
    points["Site Name"] = points["Water Point Name"].astype(str)
    points["Temporary Site"] = False
    points["Source Node"] = points["Water Point Node ID"].astype(str).map(node_index)
    assert points["Source Node"].notna().all()
    points["Source Node"] = points["Source Node"].astype(np.int32)
    points["Site Connector Distance (m)"] = pd.to_numeric(
        points["Network Snap Distance (m)"], errors="raise"
    )
    return gpd.GeoDataFrame(points, geometry="Geometry", crs=water.crs)


def construct_table() -> tuple[pd.DataFrame, dict[str, float]]:
    parameters = read("emergency_water_scenario_parameters_preprocessed.parquet")
    nodes = gpd.read_parquet(
        PROCESSED / "kumamoto_routable_road_nodes_preprocessed.parquet",
        columns=["Network Node ID", "Geometry"],
    )
    edges = read(
        "kumamoto_routable_road_edges_preprocessed.parquet",
        [
            "Road Edge ID", "From Node ID", "To Node ID", "Road Length (m)",
            "Baseline Edge Travel Time (min)", "Road Available",
            "Network Analysis Eligible",
        ],
    )
    length_graph, time_graph, node_index, retained = build_graphs(nodes, edges, set())

    water = gpd.read_parquet(
        PROCESSED / "emergency_water_points_network_access_preprocessed.parquet"
    )
    sites = build_point_inventory(water, node_index)
    demand = gpd.read_parquet(
        PROCESSED / "population_mesh_outage_demand_scenarios_preprocessed.parquet",
        columns=[
            "Mesh Code", "Geometry", "Outage Population Scenario", "Demand Scenario",
            "Estimated Outage Population", "Estimated Water Demand (L/day)",
        ],
    )
    access = read("kumamoto_population_mesh_network_access_preprocessed.parquet")
    units = demand_units(
        demand, access, retained, pd.Series(dtype="float64")
    )
    arcs = distance_arcs(
        length_graph, sites, units, access_distance_m=ACCESS_DISTANCE_M
    )
    nearest = (
        arcs.sort_values(["Unit Index", "Distance (m)", "Site Index"], kind="stable")
        .drop_duplicates("Unit Index", keep="first")
        .reset_index(drop=True)
    )

    unit_index = nearest["Unit Index"].to_numpy(np.int32)
    site_index = nearest["Site Index"].to_numpy(np.int32)
    assigned_population = np.bincount(
        site_index,
        weights=units["Estimated Outage Population"].to_numpy(float)[unit_index],
        minlength=len(sites),
    )
    assigned_water = np.bincount(
        site_index,
        weights=units["Estimated Water Demand (L/day)"].to_numpy(float)[unit_index],
        minlength=len(sites),
    )

    (
        route_site,
        route_refill,
        route_cycle,
        refill_base_minutes,
        tanker_capacity,
        trip_limit,
        work_minutes,
        _refills,
        _bases,
    ) = route_productivity(time_graph, sites, nodes, retained, parameters)
    assert tanker_capacity == 3_000
    assert trip_limit == 5
    assert work_minutes == 600

    rows: list[dict[str, object]] = []
    for point_index, point in sites.iterrows():
        required_water = float(assigned_water[point_index])
        if required_water <= 0:
            continue
        open_minutes = operating_minutes(point["Opening Time"], point["Closing Time"])
        effective_work_minutes = min(work_minutes, open_minutes)
        candidate_routes = np.flatnonzero(route_site == point_index)
        assert len(candidate_routes) > 0
        feasible: list[tuple[int, float, int]] = []
        for route_row in candidate_routes:
            refill_row = int(route_refill[route_row])
            available = effective_work_minutes - refill_base_minutes[refill_row]
            cycle = float(route_cycle[route_row])
            trips = min(trip_limit, max(0, math.floor(available / cycle)))
            feasible.append((trips, cycle, refill_row))
        best_trips, best_cycle, _ = sorted(
            feasible, key=lambda item: (-item[0], item[1], item[2])
        )[0]
        assert best_trips > 0
        assert best_trips == 5, "The displayed 15,000 L tanker-day denominator changed"
        deliverable = tanker_capacity * best_trips
        municipality_code = MUNICIPALITY_CODE_BY_JAPANESE_NAME[str(point["Municipality"])]
        rows.append(
            {
                "Municipality": MUNICIPALITY_NAMES[municipality_code],
                "Water Point": translate_exact(
                    point["Water Point Name"], ENGLISH_SITE_NAMES, "water point"
                ),
                "Operating Window": f"{point['Opening Time']}-{point['Closing Time']}",
                "Assigned Affected Residents": assigned_population[point_index],
                "Required Water (L/day)": required_water,
                "Best-Route Cycle (min/trip)": best_cycle,
                "Required Daily Trips": math.ceil(required_water / tanker_capacity),
                "Tanker-Day Capacity Required (15,000 L/day)": required_water / deliverable,
            }
        )

    table_data = pd.DataFrame(rows, columns=HEADERS).sort_values(
        ["Tanker-Day Capacity Required (15,000 L/day)", "Required Water (L/day)", "Municipality", "Water Point"],
        ascending=[False, False, True, True],
        kind="stable",
    ).reset_index(drop=True)
    total_water = float(units["Estimated Water Demand (L/day)"].sum())
    assigned_total = float(table_data["Required Water (L/day)"].sum())
    audit = {
        "all_points": float(len(sites)),
        "assigned_points": float(len(table_data)),
        "total_water": total_water,
        "assigned_water": assigned_total,
        "unassigned_water": total_water - assigned_total,
    }
    return table_data, audit


def validate_table(table_data: pd.DataFrame, audit: dict[str, float]) -> None:
    assert table_data.columns.tolist() == HEADERS
    assert 0 < len(table_data) <= 36
    assert table_data["Required Water (L/day)"].gt(0).all()
    assert table_data["Assigned Affected Residents"].gt(0).all()
    expected_trips = np.ceil(table_data["Required Water (L/day)"] / 3_000)
    assert np.array_equal(expected_trips.astype(int), table_data["Required Daily Trips"])
    expected_capacity = table_data["Required Water (L/day)"] / 15_000
    assert np.allclose(
        expected_capacity,
        table_data["Tanker-Day Capacity Required (15,000 L/day)"],
    )
    assert table_data["Tanker-Day Capacity Required (15,000 L/day)"].between(0, 1).all()
    assert np.isclose(audit["total_water"], 280_470.0, atol=1.0)
    assert audit["assigned_water"] <= audit["total_water"] + 1e-6
    for column in table_data.select_dtypes(include=["object", "string"]).columns:
        japanese_cells = table_data[column].astype(str).map(
            lambda value: bool(JAPANESE_CHARACTER_PATTERN.search(value))
        )
        assert not japanese_cells.any(), f"Japanese text remains in {column}"


def write_workbook(table_data: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False

    # Preserve the title in row 1 without merged cells for deterministic DOCX export.
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 9):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 58

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(table_data.iterrows()):
        excel_row = 3 + offset
        fill = PALE_BLUE if offset % 2 == 0 else WHITE
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(excel_row, column, record[header])
            cell.font = Font(
                name="Aptos", size=9.2,
                bold=column in (1, 2),
                color=NAVY if column in (1, 2) else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column in (1, 2, 3) else "right",
                vertical="center", wrap_text=column in (1, 2, 3),
            )
        for column in (4, 5, 7):
            sheet.cell(excel_row, column).number_format = "#,##0"
        sheet.cell(excel_row, 6).number_format = "0.0"
        sheet.cell(excel_row, 8).number_format = "0.000"
        sheet.row_dimensions[excel_row].height = 36

    end_row = len(table_data) + 2
    excel_table = Table(displayName="WaterPointTankerRequirements", ref=f"A2:H{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)
    sheet.conditional_formatting.add(
        f"H3:H{end_row}",
        ColorScaleRule(
            start_type="min", start_color="D9EAD3",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="F4CCCC",
        ),
    )
    widths = {
        "A": 23, "B": 31, "C": 20, "D": 25,
        "E": 23, "F": 25, "G": 22, "H": 31,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "C3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:H{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · 2,000 m assignment diagnostic · Required capacity, not observed supply"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY
    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d announced water-point tanker requirement scenario"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = "Kumamoto, emergency water, water point, tanker requirement"
    workbook.save(OUTPUT)


def preview_rows(table_data: pd.DataFrame) -> list[list[str]]:
    integer_columns = {
        "Assigned Affected Residents", "Required Water (L/day)",
        "Required Daily Trips",
    }
    rows: list[list[str]] = []
    for _, record in table_data.iterrows():
        row: list[str] = []
        for header in HEADERS:
            value = record[header]
            if header in integer_columns:
                text = f"{float(value):,.0f}"
            elif header == "Best-Route Cycle (min/trip)":
                text = f"{float(value):.1f}"
            elif header == "Tanker-Day Capacity Required (15,000 L/day)":
                text = f"{float(value):.3f}"
            else:
                text = str(value)
            row.append(textwrap.fill(text, width=24, break_long_words=False))
        rows.append(row)
    return rows


def write_preview(table_data: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    height = max(10.0, 2.0 + 0.48 * len(table_data))
    title_fraction = min(0.10, 0.85 / height)
    table_top = 1.0 - title_fraction
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 7})
    figure, axis = plt.subplots(figsize=(26, height))
    axis.axis("off")
    axis.add_patch(
        Rectangle(
            (0, table_top), 1, title_fraction,
            transform=axis.transAxes,
            facecolor=f"#{NAVY}", edgecolor=f"#{NAVY}", clip_on=False,
        )
    )
    axis.text(
        0.012, table_top + title_fraction / 2, TITLE,
        transform=axis.transAxes,
        ha="left", va="center",
        fontsize=18, fontweight="bold", color="white",
    )
    column_widths = [0.11, 0.18, 0.10, 0.13, 0.13, 0.12, 0.10, 0.13]
    preview = axis.table(
        cellText=preview_rows(table_data),
        colLabels=PREVIEW_HEADERS,
        colWidths=column_widths,
        cellLoc="left", colLoc="center", bbox=[0, 0, 1, table_top],
    )
    preview.auto_set_font_size(False)
    preview.set_fontsize(7.6)
    header_height = min(0.09, 0.78 / height)
    body_height = (table_top - header_height) / len(table_data)
    for (row, column), cell in preview.get_celld().items():
        cell.set_edgecolor(f"#{LIGHT_GREY}")
        cell.set_linewidth(0.45)
        if row == 0:
            cell.set_facecolor(f"#{TEAL}")
            cell.set_text_props(color="white", weight="bold", ha="center", va="center")
            cell.set_height(header_height)
        else:
            cell.set_facecolor(f"#{PALE_BLUE if (row - 1) % 2 == 0 else WHITE}")
            cell.set_text_props(
                color=f"#{NAVY}" if column in (0, 1) else f"#{TEXT}",
                weight="bold" if column in (0, 1) else "normal",
                ha="left" if column in (0, 1, 2) else "right", va="center",
            )
            cell.set_height(body_height)
        cell.PAD = 0.025
    figure.savefig(PREVIEW, dpi=320, bbox_inches="tight", pad_inches=0.06)
    plt.close(figure)


def verify_outputs(row_count: int) -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == row_count + 2 and sheet.max_column == 8
    assert sheet["A1"].value == TITLE
    assert not sheet.merged_cells.ranges
    assert sheet["A2"].value == HEADERS[0]
    assert sheet["H2"].value == HEADERS[-1]
    assert sum(
        1 for row in sheet.iter_rows(min_row=3, max_row=row_count + 2) if row[0].value
    ) == row_count
    japanese_cells = [
        cell.coordinate
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
        and JAPANESE_CHARACTER_PATTERN.search(cell.value)
    ]
    assert not japanese_cells, japanese_cells
    assert PREVIEW.exists() and PREVIEW.stat().st_size > 0


def main() -> None:
    table_data, audit = construct_table()
    validate_table(table_data, audit)
    write_workbook(table_data)
    write_preview(table_data)
    verify_outputs(len(table_data))
    print(f"Saved {len(table_data)} rows x {len(HEADERS)} cols -> {OUTPUT.relative_to(ROOT)}")
    print(f"Saved faithful PNG preview -> {PREVIEW.relative_to(ROOT)}")
    print(
        f"Assigned {audit['assigned_water']:,.0f}/{audit['total_water']:,.0f} L/day "
        f"within 2,000 m to {int(audit['assigned_points'])}/{int(audit['all_points'])} points; "
        f"unassigned {audit['unassigned_water']:,.0f} L/day"
    )


if __name__ == "__main__":
    main()
