#!/usr/bin/env python3
"""Generate Scenario-Based Priority Deployment List.

Plan: Rank the 20 selected distribution locations protecting the greatest
resident minimum-water demand in the 2,000 m extended allocation diagnostic.
Framework: AnaSOP Sections 5-7 central affected population, minimum demand of
3 L/person/day, reported water-point state, no modeled road closures, pooled
10-tanker fleet, 3,000 L loads, five-trip limit, 10-hour work limit, and the
lexicographic resident allocation model. Historical refill facilities are
planning candidates rather than verified 2026 operating sources. The 2,000 m
catchment is an allocation diagnostic, not a walking standard.
"""

from __future__ import annotations

import json
from pathlib import Path
import re
import textwrap

import geopandas as gpd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

import figure_scenario_based_tanker_and_temporary_water_point_allocation as allocation
from table_municipality_outage_population_and_water_demand import MUNICIPALITY_NAMES


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "data/results/tables/Table_scenario_based_priority_deployment_list.xlsx"
PREVIEW = ROOT / "data/exp/Table_scenario_based_priority_deployment_list.png"

TITLE = "Scenario-Based Priority Deployment List"
SHEET_NAME = "Priority Deployment"
ACCESS_DISTANCE_M = 2_000.0
SCENARIO_LABEL = "2,000 m: Capacity-constrained"
TOP_N = 20
HEADERS = [
    "Priority Rank",
    "Deployment Site",
    "Municipality",
    "Protected Residents",
    "Delivered Minimum Water (L/day)",
    "Daily Trips",
    "Mean Assigned Access (m)",
    "Historical Refill Candidate",
    "Supporting Dispatch Base",
]
PREVIEW_HEADERS = [
    "Priority\nRank",
    "Deployment Site",
    "Municipality",
    "Protected\nResidents",
    "Delivered Minimum\nWater (L/day)",
    "Daily\nTrips",
    "Mean Assigned\nAccess (m)",
    "Historical Refill\nCandidate",
    "Supporting Dispatch\nBase",
]

NAVY = "17365D"
TEAL = "0B6E75"
PALE_BLUE = "EAF0F8"
WHITE = "FFFFFF"
TEXT = "1D2939"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"

ENGLISH_SITE_NAMES = {
    "二見コミセン": "Futami Community Center",
    "八代小学校": "Yatsushiro Elementary School",
    "松高コミセン": "Matsutaka Community Center",
    "第七中学校": "Yatsushiro Seventh Junior High School",
    "第四中学校": "Yatsushiro Fourth Junior High School",
    "麦島コミセン": "Mugishima Community Center",
    "第五中学校": "Yatsushiro Fifth Junior High School",
    "日奈久小学校": "Hinagu Elementary School",
    "日奈久コミセン": "Hinagu Community Center",
    "昭和コミセン": "Showa Community Center",
    "八代支援学校": "Yatsushiro Special Needs School",
    "金剛小学校": "Kongo Elementary School",
    "金剛コミセン": "Kongo Community Center",
    "植柳小学校": "Uyanagi Elementary School",
    "松高小学校": "Matsutaka Elementary School",
    "宮地コミセン": "Miyaji Community Center",
    "太田郷小学校": "Otagō Elementary School",
    "第一中学校": "Yatsushiro First Junior High School",
    "文政小学校": "Bunsei Elementary School",
    "市役所本庁舎西側": "West Side of Yatsushiro City Hall",
    "第二中学校": "Yatsushiro Second Junior High School",
    "代陽コミセン": "Daiyo Community Center",
    "八千把コミセン": "Yachiwa Community Center",
    "太田郷コミセン": "Otagō Community Center",
    "鏡コミセン": "Kagami Community Center",
    "鏡町野崎公民館": "Nozaki Public Hall, Kagami Town",
    "千丁コミセン": "Sencho Community Center",
    "不知火防災拠点センター": "Shiranui Disaster Management Center",
    "宇城市役所": "Uki City Hall",
    "松橋東防災拠点センター": "Matsubase East Disaster Management Center",
    "松橋西防災拠点センター": "Matsubase West Disaster Management Center",
    "小川防災拠点センター": "Ogawa Disaster Management Center",
    "豊野防災拠点センター": "Toyono Disaster Management Center",
    "氷川町役場": "Hikawa Town Hall",
    "宮原振興局": "Miyahara Regional Office",
    "龍峯コミセン": "Ryuho Community Center",
}

ENGLISH_REFILL_NAMES = {
    "片岩浄水場": "Kataiwa Water Treatment Plant",
    "小崎浄水場": "Kozaki Water Treatment Plant",
    "東部第２浄水場": "Tobu No. 2 Water Treatment Plant",
    "八代浄水場": "Yatsushiro Water Treatment Plant",
    "豊川浄水場": "Toyokawa Water Treatment Plant",
    "花園浄水場": "Hanazono Water Treatment Plant",
    "小川浄水場": "Ogawa Water Treatment Plant",
}

ENGLISH_BASE_NAMES = {
    "八代広域行政事務組合八代消防署坂本分署": (
        "Yatsushiro Fire Station, Sakamoto Branch"
    ),
    "八代広域行政事務組合八代消防署新開分署": (
        "Yatsushiro Fire Station, Shinkai Branch"
    ),
    "宇城広域連合北消防署": "Uki Regional North Fire Station",
    "宇城広域連合南消防署": "Uki Regional South Fire Station",
}

JAPANESE_CHARACTER_PATTERN = re.compile(r"[\u3040-\u30ff\u3400-\u9fff]")


def translate_exact(value: object, mapping: dict[str, str], field: str) -> str:
    source = str(value)
    if source not in mapping:
        raise ValueError(f"Missing English translation for {field}: {source}")
    return mapping[source]


def english_refill_label(refill: pd.Series) -> str:
    source = str(refill["Water Treatment Facility Name"])
    if source in ENGLISH_REFILL_NAMES:
        return ENGLISH_REFILL_NAMES[source]
    identifier = str(refill["P21 Inspection ID"])
    if JAPANESE_CHARACTER_PATTERN.search(identifier):
        raise ValueError("Historical refill identifier contains Japanese text")
    return f"Historical Refill Candidate {identifier}"


def english_base_label(base: pd.Series) -> str:
    source = str(base["Fire Facility Name"])
    if source in ENGLISH_BASE_NAMES:
        return ENGLISH_BASE_NAMES[source]
    identifier = str(base["Dispatch Base Node ID"])
    if JAPANESE_CHARACTER_PATTERN.search(identifier):
        raise ValueError("Dispatch-base identifier contains Japanese text")
    return f"Dispatch Base {identifier}"


def read_inputs() -> tuple[
    pd.DataFrame,
    gpd.GeoDataFrame,
    pd.DataFrame,
    gpd.GeoDataFrame,
    pd.DataFrame,
    gpd.GeoDataFrame,
    gpd.GeoDataFrame,
]:
    parameters = pd.read_parquet(allocation.PARAMETERS)
    nodes = gpd.read_parquet(
        allocation.ROAD_NODES, columns=["Network Node ID", "Geometry"]
    )
    edges = pd.read_parquet(
        allocation.ROAD_EDGES,
        columns=[
            "Road Edge ID",
            "From Node ID",
            "To Node ID",
            "Road Length (m)",
            "Baseline Edge Travel Time (min)",
            "Road Available",
            "Network Analysis Eligible",
        ],
    )
    demand = gpd.read_parquet(allocation.MESH_DEMAND)
    access = pd.read_parquet(allocation.MESH_ACCESS)
    water = gpd.read_parquet(allocation.WATER_POINTS)
    staging = gpd.read_parquet(allocation.STAGING)
    return parameters, nodes, edges, demand, access, water, staging


def reconstruct_audited_extended_scenario() -> tuple[
    allocation.ScenarioInputs, allocation.AllocationResult, int
]:
    """Rebuild the accepted 2,000 m ledger from the figure's rerun audit.

    The allocation figure solves every scenario twice.  This table consumes the
    exact-agreement signature from that audit so its ranked locations cannot
    drift from the accepted map because of another near-optimal MILP run.
    """
    parameters, nodes, edges, demand, access, water, staging = read_inputs()
    fleet_size = int(allocation.parameter_values(parameters, "Fleet Size")["central"])
    assert fleet_size == 10
    scenario = allocation.prepare_scenario(
        SCENARIO_LABEL,
        set(),
        None,
        nodes,
        edges,
        demand,
        access,
        water,
        staging,
        parameters,
        allocation.older_priority_weights(),
        access_distance_m=ACCESS_DISTANCE_M,
    )

    audit_path = allocation.SOLVER_AUDIT_PATH
    if not audit_path.exists():
        raise FileNotFoundError(
            "Run the allocation-figure script first to create the solver audit"
        )
    solver_audit = json.loads(audit_path.read_text(encoding="utf-8"))
    reproducibility = next(
        (
            item
            for item in solver_audit["reproducibility"]
            if item["scenario"] == SCENARIO_LABEL
        ),
        None,
    )
    if reproducibility is None:
        raise ValueError(f"Solver audit has no {SCENARIO_LABEL} result")
    if not reproducibility["exact_solution_agreement"]:
        raise ValueError(
            "The accepted 2,000 m allocation lacks exact independent-rerun agreement"
        )
    signature = reproducibility["first_signature"]
    if signature != reproducibility["repeat_signature"]:
        raise ValueError("Independent-rerun signatures disagree")
    run_audit = next(
        item for item in solver_audit["runs"] if item["scenario"] == SCENARIO_LABEL
    )

    unit_lookup = {
        str(mesh_code): int(index)
        for index, mesh_code in scenario.units["Mesh Code"].items()
    }
    site_lookup = {
        str(site_id): int(index)
        for index, site_id in scenario.sites["Site ID"].items()
    }
    assignment_rows = []
    for mesh_code, site_id in signature["assignments"]:
        if str(mesh_code) not in unit_lookup or str(site_id) not in site_lookup:
            raise ValueError("Audited assignment cannot be linked to the current scenario")
        assignment_rows.append(
            {
                "Unit Index": unit_lookup[str(mesh_code)],
                "Site Index": site_lookup[str(site_id)],
            }
        )
    assignment_index = pd.DataFrame(assignment_rows)
    selected_arcs = assignment_index.merge(
        scenario.arcs,
        on=["Unit Index", "Site Index"],
        how="left",
        validate="one_to_one",
    )
    if selected_arcs["Distance (m)"].isna().any():
        raise ValueError("Audited assignment is not eligible in the current scenario")

    route_trips = np.zeros(len(scenario.route_site_index), dtype=int)
    refill_ids = scenario.refills["P21 Inspection ID"].astype(str).to_numpy()
    for site_id, refill_id, trips in signature["route_plan"]:
        site_index = site_lookup[str(site_id)]
        route_rows = np.flatnonzero(
            (scenario.route_site_index == site_index)
            & (refill_ids[scenario.route_refill_index] == str(refill_id))
        )
        if len(route_rows) != 1:
            raise ValueError("Audited route is not unique in the current scenario")
        route_trips[int(route_rows[0])] = int(trips)

    vehicles = np.zeros(len(scenario.refills), dtype=int)
    for refill_id, vehicle_count in signature["vehicle_plan"]:
        refill_rows = np.flatnonzero(refill_ids == str(refill_id))
        if len(refill_rows) != 1:
            raise ValueError("Audited refill candidate is not unique")
        vehicles[int(refill_rows[0])] = int(vehicle_count)

    trips_by_site = np.bincount(
        scenario.route_site_index,
        weights=route_trips,
        minlength=len(scenario.sites),
    ).astype(int)
    selected_indices = np.flatnonzero(trips_by_site > 0)
    selected_sites = scenario.sites.loc[selected_indices].copy()
    selected_sites["Trips"] = trips_by_site[selected_indices]
    selected_sites["Site Index"] = selected_indices
    selected_sites["Best Refill Row"] = -1
    for site_index in selected_indices:
        route_rows = np.flatnonzero(
            (scenario.route_site_index == site_index) & (route_trips > 0)
        )
        best_route = route_rows[
            np.lexsort(
                (
                    scenario.route_cycle_minutes[route_rows],
                    -route_trips[route_rows],
                )
            )[0]
        ]
        selected_sites.loc[
            selected_sites["Site Index"].eq(site_index), "Best Refill Row"
        ] = int(scenario.route_refill_index[best_route])
    selected_sites["Best Refill Row"] = selected_sites["Best Refill Row"].astype(int)

    delivered = float(
        scenario.units.iloc[
            selected_arcs["Unit Index"].to_numpy(np.int32)
        ]["Estimated Water Demand (L/day)"].sum()
    )
    result = allocation.AllocationResult(
        delivery_liters=delivered,
        route_trips=route_trips,
        trips_by_site=trips_by_site,
        vehicles_by_refill=vehicles,
        selected_arcs=selected_arcs,
        selected_sites=selected_sites,
        failed_point=None,
        solver_audit=tuple(run_audit["stages"]),
    )
    allocation.validate_solution(scenario, result, fleet_size)
    reconstructed_signature = json.loads(
        json.dumps(allocation.allocation_signature(scenario, result), sort_keys=True)
    )
    if reconstructed_signature != signature:
        raise ValueError("Reconstructed table ledger differs from the accepted figure")
    return scenario, result, fleet_size


def site_municipalities(sites: gpd.GeoDataFrame) -> pd.Series:
    municipalities = gpd.read_parquet(allocation.MUNICIPALITIES)[
        ["Reporting Municipality Code", "Geometry"]
    ]
    points = sites[["Site Index", "Geometry"]].copy()
    if points.crs != municipalities.crs:
        points = points.to_crs(municipalities.crs)
    joined = gpd.sjoin(points, municipalities, how="left", predicate="within")
    codes = joined.drop_duplicates("Site Index").set_index("Site Index")[
        "Reporting Municipality Code"
    ]
    return codes.astype("string").map(MUNICIPALITY_NAMES)


def construct_table() -> tuple[pd.DataFrame, dict[str, float]]:
    scenario, result, fleet_size = reconstruct_audited_extended_scenario()
    units = scenario.units
    selected_arcs = result.selected_arcs.copy()
    unit_index = selected_arcs["Unit Index"].to_numpy(np.int32)
    site_index = selected_arcs["Site Index"].to_numpy(np.int32)
    population = units["Estimated Outage Population"].to_numpy(float)[unit_index]
    water = units["Estimated Water Demand (L/day)"].to_numpy(float)[unit_index]
    distance = selected_arcs["Distance (m)"].to_numpy(float)

    selected_arcs["Protected Residents"] = population
    selected_arcs["Delivered Minimum Water (L/day)"] = water
    selected_arcs["Population-Distance"] = population * distance
    aggregates = selected_arcs.groupby("Site Index", observed=True).agg(
        **{
            "Protected Residents": ("Protected Residents", "sum"),
            "Delivered Minimum Water (L/day)": (
                "Delivered Minimum Water (L/day)", "sum"
            ),
            "Population-Distance": ("Population-Distance", "sum"),
        }
    )
    aggregates["Mean Assigned Access (m)"] = (
        aggregates["Population-Distance"] / aggregates["Protected Residents"]
    )

    selected_sites = result.selected_sites.set_index("Site Index").copy()
    selected_sites["Site Name"] = selected_sites["Site Name"].map(
        lambda value: translate_exact(value, ENGLISH_SITE_NAMES, "deployment site")
    )
    selected_sites["Municipality"] = site_municipalities(result.selected_sites)
    selected_sites = selected_sites.join(aggregates, how="inner", validate="one_to_one")
    selected_sites["Daily Trips"] = result.trips_by_site[
        selected_sites.index.to_numpy(np.int32)
    ]

    refill_names: list[str] = []
    base_names: list[str] = []
    for _, site in selected_sites.iterrows():
        refill_row = int(site["Best Refill Row"])
        refill = scenario.refills.iloc[refill_row]
        base = scenario.dispatch_bases.iloc[int(refill["Supporting Base Row"])]
        refill_names.append(english_refill_label(refill))
        base_names.append(english_base_label(base))
    selected_sites["Historical Refill Candidate"] = refill_names
    selected_sites["Supporting Dispatch Base"] = base_names

    assert not selected_sites["Temporary Site"].any(), (
        "The current 2,000 m solution now selects a temporary site; revise the "
        "publication table to report site type before continuing."
    )
    ranked = selected_sites.sort_values(
        ["Delivered Minimum Water (L/day)", "Protected Residents", "Site Name"],
        ascending=[False, False, True],
        kind="stable",
    ).head(TOP_N).copy()
    ranked.insert(0, "Priority Rank", np.arange(1, len(ranked) + 1))
    ranked = ranked.rename(columns={"Site Name": "Deployment Site"})
    table_data = ranked[
        [
            "Priority Rank",
            "Deployment Site",
            "Municipality",
            "Protected Residents",
            "Delivered Minimum Water (L/day)",
            "Daily Trips",
            "Mean Assigned Access (m)",
            "Historical Refill Candidate",
            "Supporting Dispatch Base",
        ]
    ].reset_index(drop=True)

    audit = {
        "fleet_size": float(fleet_size),
        "fleet_used": float(result.fleet_used),
        "trips_used": float(result.trips_used),
        "selected_sites": float(len(result.selected_sites)),
        "temporary_sites": float(result.selected_sites["Temporary Site"].sum()),
        "total_demand": float(units["Estimated Water Demand (L/day)"].sum()),
        "delivered_water": float(result.delivery_liters),
        "top_n_water": float(table_data["Delivered Minimum Water (L/day)"].sum()),
    }
    return table_data, audit


def validate_table(table_data: pd.DataFrame, audit: dict[str, float]) -> None:
    assert table_data.columns.tolist() == HEADERS
    assert table_data.shape == (TOP_N, len(HEADERS))
    assert table_data["Priority Rank"].tolist() == list(range(1, TOP_N + 1))
    assert table_data["Municipality"].notna().all()
    assert table_data["Protected Residents"].gt(0).all()
    assert table_data["Delivered Minimum Water (L/day)"].gt(0).all()
    assert table_data["Daily Trips"].between(1, 5).all()
    assert table_data["Mean Assigned Access (m)"].between(0, ACCESS_DISTANCE_M).all()
    assert np.allclose(
        table_data["Delivered Minimum Water (L/day)"],
        3.0 * table_data["Protected Residents"],
    )
    assert table_data["Delivered Minimum Water (L/day)"].is_monotonic_decreasing
    for column in table_data.select_dtypes(include=["object", "string"]).columns:
        japanese_cells = table_data[column].astype(str).map(
            lambda value: bool(JAPANESE_CHARACTER_PATTERN.search(value))
        )
        assert not japanese_cells.any(), f"Japanese text remains in {column}"
    assert audit["fleet_size"] == 10
    assert audit["fleet_used"] <= audit["fleet_size"]
    assert audit["trips_used"] <= 50
    assert audit["selected_sites"] == 28
    assert audit["temporary_sites"] == 0
    assert np.isclose(audit["total_demand"], 280_470.0, atol=1.0)
    assert np.isclose(audit["delivered_water"], 148_745.67038000844, atol=1e-4)
    assert audit["top_n_water"] <= audit["delivered_water"] + 1e-6


def write_workbook(table_data: pd.DataFrame) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False

    # Preserve the title in row 1 without merged cells for deterministic DOCX export.
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 10):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 54

    thin = Side(style="thin", color=LIGHT_GREY)
    for offset, (_, record) in enumerate(table_data.iterrows()):
        excel_row = 3 + offset
        fill = PALE_BLUE if offset % 2 == 0 else WHITE
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(excel_row, column, record[header])
            left_aligned = column in (2, 3, 8, 9)
            cell.font = Font(
                name="Aptos",
                size=9.3,
                bold=column in (1, 2),
                color=NAVY if column in (1, 2) else TEXT,
            )
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if left_aligned else "right",
                vertical="center",
                wrap_text=left_aligned,
            )
        for column in (1, 4, 5, 6):
            sheet.cell(excel_row, column).number_format = "#,##0"
        sheet.cell(excel_row, 7).number_format = "#,##0.0"
        sheet.row_dimensions[excel_row].height = 34

    end_row = len(table_data) + 2
    excel_table = Table(displayName="PriorityDeploymentList", ref=f"A2:I{end_row}")
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)
    sheet.conditional_formatting.add(
        f"E3:E{end_row}",
        ColorScaleRule(
            start_type="min",
            start_color="D9EAD3",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="F4CCCC",
        ),
    )
    widths = {
        "A": 13,
        "B": 30,
        "C": 20,
        "D": 21,
        "E": 27,
        "F": 14,
        "G": 24,
        "H": 27,
        "I": 29,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "D3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:I{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = (
        "KE01d · 2,000 m extended allocation diagnostic · Central/minimum demand · 10-tanker fleet"
    )
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY
    workbook.properties.title = TITLE
    workbook.properties.subject = "KE01d optimized priority deployment scenario"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = (
        "Kumamoto, emergency water, tanker allocation, priority deployment"
    )
    workbook.save(OUTPUT)


def preview_rows(table_data: pd.DataFrame) -> list[list[str]]:
    integer_columns = {
        "Priority Rank",
        "Protected Residents",
        "Delivered Minimum Water (L/day)",
        "Daily Trips",
    }
    rows: list[list[str]] = []
    for _, record in table_data.iterrows():
        row: list[str] = []
        for header in HEADERS:
            value = record[header]
            if header in integer_columns:
                value_text = f"{float(value):,.0f}"
            elif header == "Mean Assigned Access (m)":
                value_text = f"{float(value):,.1f}"
            else:
                value_text = str(value)
            row.append(textwrap.fill(value_text, width=25, break_long_words=False))
        rows.append(row)
    return rows


def write_preview(table_data: pd.DataFrame) -> None:
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    height = max(9.0, 2.1 + 0.48 * len(table_data))
    title_fraction = min(0.10, 0.86 / height)
    table_top = 1.0 - title_fraction
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 7.2})
    figure, axis = plt.subplots(figsize=(27, height))
    axis.axis("off")
    axis.add_patch(
        Rectangle(
            (0, table_top),
            1,
            title_fraction,
            transform=axis.transAxes,
            facecolor=f"#{NAVY}",
            edgecolor=f"#{NAVY}",
            clip_on=False,
        )
    )
    axis.text(
        0.012,
        table_top + title_fraction / 2,
        TITLE,
        transform=axis.transAxes,
        ha="left",
        va="center",
        fontsize=18,
        fontweight="bold",
        color="white",
    )
    column_widths = [0.055, 0.17, 0.095, 0.105, 0.125, 0.065, 0.115, 0.125, 0.145]
    preview = axis.table(
        cellText=preview_rows(table_data),
        colLabels=PREVIEW_HEADERS,
        colWidths=column_widths,
        cellLoc="left",
        colLoc="center",
        bbox=[0, 0, 1, table_top],
    )
    preview.auto_set_font_size(False)
    preview.set_fontsize(7.8)
    header_height = min(0.10, 0.80 / height)
    body_height = (table_top - header_height) / len(table_data)
    for (row, column), cell in preview.get_celld().items():
        cell.set_edgecolor(f"#{LIGHT_GREY}")
        cell.set_linewidth(0.45)
        if row == 0:
            cell.set_facecolor(f"#{TEAL}")
            cell.set_text_props(color="white", weight="bold", ha="center", va="center")
            cell.set_height(header_height)
        else:
            cell.set_facecolor(f"#{PALE_BLUE if (row - 1) % 2 == 0 else WHITE}")
            left_aligned = column in (1, 2, 7, 8)
            cell.set_text_props(
                color=f"#{NAVY}" if column in (0, 1) else f"#{TEXT}",
                weight="bold" if column in (0, 1) else "normal",
                ha="left" if left_aligned else "right",
                va="center",
            )
            cell.set_height(body_height)
        cell.PAD = 0.025
    figure.savefig(PREVIEW, dpi=320, bbox_inches="tight", pad_inches=0.06)
    plt.close(figure)


def verify_outputs(row_count: int) -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook[SHEET_NAME]
    assert sheet.max_row == row_count + 2 and sheet.max_column == len(HEADERS)
    assert sheet["A1"].value == TITLE
    assert not sheet.merged_cells.ranges
    assert sheet["A2"].value == HEADERS[0]
    assert sheet["I2"].value == HEADERS[-1]
    assert sum(
        1 for row in sheet.iter_rows(min_row=3, max_row=row_count + 2) if row[0].value
    ) == row_count
    assert PREVIEW.exists() and PREVIEW.stat().st_size > 0


def main() -> None:
    table_data, audit = construct_table()
    validate_table(table_data, audit)
    write_workbook(table_data)
    write_preview(table_data)
    verify_outputs(len(table_data))
    print(f"Saved {len(table_data)} rows x {len(HEADERS)} cols -> {OUTPUT.relative_to(ROOT)}")
    print(f"Saved faithful PNG preview -> {PREVIEW.relative_to(ROOT)}")
    print(
        f"Scenario: {int(audit['fleet_used'])}/{int(audit['fleet_size'])} tanker equivalents; "
        f"{int(audit['trips_used'])} trips; {int(audit['selected_sites'])} selected points; "
        f"{int(audit['temporary_sites'])} temporary; "
        f"delivered {audit['delivered_water']:,.0f}/{audit['total_demand']:,.0f} L/day"
    )


if __name__ == "__main__":
    main()
