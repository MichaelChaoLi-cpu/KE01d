#!/usr/bin/env python3
"""Generate the planned Data Linkage and Coverage Audit workbook.

The table recomputes every count from analysis-ready parquet files. Its nine
rows and seven columns follow docs/AnaSOP.md Section 8. Coverage measures data
linkage only; it must not be interpreted as evidence of operational capacity.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
OUTPUT = ROOT / "data/results/tables/Table_data_linkage_and_coverage_audit.xlsx"

NAVY = "17365D"
TEAL = "0B6E75"
PALE_TEAL = "E8F3F3"
PALE_BLUE = "EAF0F8"
PALE_ORANGE = "FCE4D6"
WHITE = "FFFFFF"
MID_GREY = "667085"
LIGHT_GREY = "D0D5DD"


def read(name: str) -> pd.DataFrame:
    return pd.read_parquet(PROCESSED / name)


def audit_rows() -> list[dict[str, object]]:
    crosswalk = read("reporting_unit_municipality_crosswalk_preprocessed.parquet")
    mesh_demand = read("population_mesh_outage_demand_scenarios_preprocessed.parquet")
    mesh_network = read("kumamoto_population_mesh_network_access_preprocessed.parquet")
    older_network = read("kumamoto_population_group_network_access_preprocessed.parquet")
    water = read("emergency_water_points_network_access_preprocessed.parquet")
    shelters = read("shelter_water_demand_network_access_preprocessed.parquet").drop_duplicates(
        "Shelter Number"
    )
    staging = read("kumamoto_staging_site_candidates_preprocessed.parquet")
    bases = read("kumamoto_dispatch_base_network_access_preprocessed.parquet")
    restrictions = read("road_restriction_edge_matches_preprocessed.parquet").drop_duplicates(
        "Restriction Observation ID"
    )

    mesh_municipality = mesh_demand.drop_duplicates("Mesh Code")
    water_success = water["Location Resolution Status"].ne("unresolved") & water[
        "Network Snap Accepted"
    ].fillna(False)
    shelter_success = shelters["Location Resolution Status"].ne("unresolved") & shelters[
        "Network Snap Accepted"
    ].fillna(False)

    specifications = [
        (
            "Reporting-unit municipality linkage",
            len(crosswalk),
            crosswalk["Municipality Match Status"].eq("exact_official_name").sum(),
            "Exact official-name crosswalk; only in-scope Kumamoto municipalities are accepted.",
            "Four non-eligible records remain visible: 3 neighboring-prefecture units and 1 joint operator area; none is allocated to a municipality.",
            "reporting_unit_municipality_crosswalk_preprocessed.parquet",
        ),
        (
            "Population-mesh municipality linkage",
            len(mesh_municipality),
            mesh_municipality["Spatial Join Status"].ne("unmatched").sum(),
            "Unique point match; maximum-area overlap is used only where a unique point match is unavailable.",
            "Six meshes remain in an explicit unmatched audit unit and are excluded from municipality-specific estimates.",
            "population_mesh_outage_demand_scenarios_preprocessed.parquet",
        ),
        (
            "Population-mesh road-network linkage",
            len(mesh_network),
            mesh_network["Network Snap Accepted"].fillna(False).sum(),
            "The inherited Network Snap Accepted flag defines a usable demand-side road node.",
            "Forty-eight rejected meshes have no route-distance result, but remain in resident-demand denominators.",
            "kumamoto_population_mesh_network_access_preprocessed.parquet",
        ),
        (
            "Older-population-group road-network linkage",
            len(older_network),
            older_network["Network Snap Accepted"].fillna(False).sum(),
            "The inherited Network Snap Accepted flag is applied at disclosure-group support.",
            "Seven rejected groups have no route-distance result; older-population totals are not copied to constituent meshes.",
            "kumamoto_population_group_network_access_preprocessed.parquet",
        ),
        (
            "Announced water-point location and road-network linkage",
            len(water),
            water_success.sum(),
            "Deterministic coordinate hierarchy; exact-name candidates require one unique geometry; accepted network snap ≤250 m.",
            "Complete linkage does not verify opening hours, capacity, water availability, or current operating status.",
            "emergency_water_points_network_access_preprocessed.parquet",
        ),
        (
            "Public-shelter location and road-network linkage",
            len(shelters),
            shelter_success.sum(),
            "Deterministic official/exact location resolution followed by an accepted network snap ≤250 m.",
            "Counts use 41 unique shelters; the demand file expands them to 123 scenario rows and does not prove evacuee-count completeness.",
            "shelter_water_demand_network_access_preprocessed.parquet",
        ),
        (
            "Candidate staging-site road-network linkage",
            len(staging),
            staging["Network Snap Accepted"].fillna(False).sum(),
            "A staging candidate succeeds when its inherited demand-node/network snap is accepted.",
            "One hundred fifty-nine sites lack network nodes; only 35 pass the inherited screen, which is not a 2026 deployment decision.",
            "kumamoto_staging_site_candidates_preprocessed.parquet",
        ),
        (
            "Historical dispatch-base road-network linkage",
            len(bases),
            (
                bases["Candidate Dispatch Base"].fillna(False)
                & bases["Network Snap Accepted"].fillna(False)
            ).sum(),
            "Candidate Dispatch Base must be true and the dispatch-base network snap must be accepted.",
            "These are historical fire-facility candidates and do not establish tanker presence, capacity, or availability.",
            "kumamoto_dispatch_base_network_access_preprocessed.parquet",
        ),
        (
            "Road-restriction observation road-edge linkage",
            len(restrictions),
            restrictions["Road Edge Match Status"].ne("unmatched").sum(),
            "Line ≤50 m or point ≤100 m; otherwise nearest-edge fallback ≤250 m; all observations are retained.",
            "A spatial edge match is a scenario input, not confirmation that the matched road was closed or impassable.",
            "road_restriction_edge_matches_preprocessed.parquet",
        ),
    ]

    rows = []
    for component, total, successful, rule, limit, source in specifications:
        total = int(total)
        successful = int(successful)
        rows.append(
            {
                "Component": component,
                "Total Records": total,
                "Successful Records": successful,
                "Unmatched Records": total - successful,
                "Coverage Rate": successful / total if total else None,
                "Acceptance Rule": rule,
                "Interpretation Limit": limit,
                "_source": source,
            }
        )
    return rows


def validate(rows: list[dict[str, object]]) -> None:
    assert len(rows) == 9, "AnaSOP requires nine audit components"
    expected = [
        (19, 15),
        (62_945, 62_939),
        (62_945, 62_897),
        (36_657, 36_650),
        (36, 36),
        (41, 41),
        (6_105, 5_946),
        (81, 81),
        (680, 604),
    ]
    observed = [(row["Total Records"], row["Successful Records"]) for row in rows]
    assert observed == expected, f"Unexpected linkage counts: {observed}"
    for row in rows:
        assert row["Unmatched Records"] == row["Total Records"] - row["Successful Records"]
        assert 0 <= row["Coverage Rate"] <= 1


def write_workbook(rows: list[dict[str, object]]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Linkage Audit"
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "Data Linkage and Coverage Audit"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    for column in range(1, 8):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    headers = [
        "Component",
        "Total Records",
        "Successful Records",
        "Unmatched Records",
        "Coverage Rate",
        "Acceptance Rule",
        "Interpretation Limit",
    ]
    header_row = 2
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 34

    thin = Side(style="thin", color=LIGHT_GREY)
    start_row = header_row + 1
    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        values = [row[header] for header in headers]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(excel_row, column, value)
            cell.font = Font(name="Aptos", size=10, color="1D2939")
            cell.fill = PatternFill(
                "solid", fgColor=PALE_BLUE if offset % 2 == 0 else WHITE
            )
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column in (1, 6, 7) else "right",
                vertical="top",
                wrap_text=column in (1, 6, 7),
            )
        sheet.cell(excel_row, 1).font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        sheet.cell(excel_row, 1).comment = Comment(
            f"Recomputed from data/processed/{row['_source']}", "MiliFrame"
        )
        for column in (2, 3, 4):
            sheet.cell(excel_row, column).number_format = "#,##0"
        sheet.cell(excel_row, 5).number_format = "0.0%"
        sheet.row_dimensions[excel_row].height = 58

    end_row = start_row + len(rows) - 1
    table = Table(displayName="DataLinkageCoverageAudit", ref=f"A{header_row}:G{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.conditional_formatting.add(
        f"D{start_row}:D{end_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=PALE_ORANGE),
        ),
    )
    sheet.conditional_formatting.add(
        f"E{start_row}:E{end_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0.75,
            start_color="F4CCCC",
            mid_type="num",
            mid_value=0.95,
            mid_color="FFF2CC",
            end_type="num",
            end_value=1.0,
            end_color="D9EAD3",
        ),
    )

    widths = {"A": 39, "B": 16, "C": 18, "D": 17, "E": 15, "F": 54, "G": 66}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B3"
    sheet.auto_filter.ref = f"A{header_row}:G{end_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = f"A1:G{end_row}"
    sheet.print_title_rows = "1:2"
    sheet.oddFooter.center.text = "KE01d · Research planning scenario · Generated from processed data"
    sheet.oddFooter.center.size = 8
    sheet.oddFooter.center.color = MID_GREY

    workbook.properties.title = "Data Linkage and Coverage Audit"
    workbook.properties.subject = "KE01d emergency-water planning linkage diagnostics"
    workbook.properties.creator = "Mike Li"
    workbook.properties.keywords = "Kumamoto, emergency water, linkage audit, MiliFrame"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(OUTPUT)


def verify_workbook() -> None:
    workbook = load_workbook(OUTPUT, data_only=False)
    sheet = workbook["Data Linkage Audit"]
    assert sheet.max_column == 7
    assert sheet.max_row == 11
    assert sheet["A2"].value == "Component"
    assert sheet["G2"].value == "Interpretation Limit"
    assert sum(1 for row in sheet.iter_rows(min_row=3, max_row=11) if row[0].value) == 9
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("#")
        for row in sheet.iter_rows()
        for cell in row
    )


def main() -> None:
    rows = audit_rows()
    validate(rows)
    write_workbook(rows)
    verify_workbook()
    print(f"Saved {len(rows)} rows x 7 cols -> {OUTPUT.relative_to(ROOT)}")
    for row in rows:
        print(
            f"{row['Component']}: {row['Successful Records']:,}/{row['Total Records']:,} "
            f"({row['Coverage Rate']:.1%})"
        )


if __name__ == "__main__":
    main()
